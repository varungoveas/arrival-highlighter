import streamlit as st
import pdfplumber
import re
import io
from pypdf import PdfReader, PdfWriter
from pypdf.generic import (ArrayObject, FloatObject, NameObject,
                            DictionaryObject, TextStringObject)
from collections import defaultdict
from datetime import date
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib import colors
from reportlab.lib.units import mm
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.styles import ParagraphStyle
from reportlab.lib.enums import TA_LEFT, TA_CENTER

st.set_page_config(page_title="Arrival Report Highlighter", page_icon="🏝️", layout="wide")
st.markdown("""
<style>
#MainMenu, footer, header {visibility: hidden;}
.main .block-container {padding-top: 1.5rem; padding-bottom: 1rem;}
h1 {font-size: 1.6rem !important; margin-bottom: 0 !important;}
</style>
""", unsafe_allow_html=True)

YELLOW = (1.0, 0.921, 0.231)
ORANGE = (1.0, 0.600, 0.100)

CATEGORIES = {
    'occasion':    {'label': 'Special Occasion',       'icon': '🎂',
                    'kw': ['birthday','anniversary','honeymoon','babymoon','celebrating','celebration','proposal','engagement','wedding']},
    'allergy':     {'label': 'Allergy / Dietary',      'icon': '🌿',
                    'kw': ['allerg','intoleranc','gluten','lactose','dietary restriction','vegan','vegetarian','halal','kosher','peanut','shellfish']},
    'payment':     {'label': 'Outstanding Payment',    'icon': '💳',
                    'kw': ['collect upon arrival','please collect','pls collect','balance due','outstanding balance','to be collected','pay upon arrival']},
    'flight':      {'label': 'Flights (🟡 ETA present / 🟠 ETA missing)', 'icon': '✈️', 'kw': []},
    'repeater':    {'label': 'Repeater Guest',         'icon': '⭐', 'kw': []},
    'stayhistory': {'label': 'Stay History',           'icon': '🏨', 'kw': []},
    'complaint':   {'label': 'Complaint / Glitch',     'icon': '⚠️',
                    'kw': ['glitch','recovery','complaint','inconvenien','apologi','disatisf','dissatisf','ttnglitch','ttncomp','feedback','upset']},
    'membership':  {'label': 'GHA Membership',         'icon': '👑', 'kw': []},
    'dbalance':    {'label': 'D$ Balance (> 0)',        'icon': '💰', 'kw': []},
    'legs':        {'label': 'Multi-Villa Legs',       'icon': '🏝️',
                    'kw': ['1st leg','2nd leg','3rd leg','4th leg','5th leg','1ST Leg','2nd Leg','leg:','leg -','leg-','leg of the stay']},
    'vip':         {'label': 'VIP Guests',             'icon': '💎', 'kw': []},
    'sharewith':   {'label': 'Travelling Together (Share with tag)', 'icon': '👥', 'kw': []},
    'welcomenote': {'label': 'Welcome Note',           'icon': '📝',
                    'kw': ['welcome note','welcome amenities','welcome fruit','welcome cake','welcome letter','welcome card','welcome drink','welcome set']},
    'earlyarr':    {'label': 'Early Arrival',          'icon': '🌅', 'kw': []},
    'roomno':      {'label': 'Room / Name / Conf No.', 'icon': '🚪', 'kw': []},
}

def make_annot(x0, top, x1, bottom, ph, color, popup=None):
    px0, py0, px1, py1 = x0-1, ph-bottom-1, x1+1, ph-top+1
    a = DictionaryObject()
    a[NameObject('/Type')]    = NameObject('/Annot')
    a[NameObject('/Subtype')] = NameObject('/Highlight')
    a[NameObject('/Rect')]    = ArrayObject([FloatObject(px0), FloatObject(py0),
                                             FloatObject(px1), FloatObject(py1)])
    a[NameObject('/C')]       = ArrayObject([FloatObject(c) for c in color])
    a[NameObject('/QuadPoints')] = ArrayObject([
        FloatObject(px0), FloatObject(py1), FloatObject(px1), FloatObject(py1),
        FloatObject(px0), FloatObject(py0), FloatObject(px1), FloatObject(py0)])
    a[NameObject('/F')] = FloatObject(4)
    if popup:
        a[NameObject('/Contents')] = TextStringObject(popup)
    return a


def merge_groups(travelling_together):
    """Merge connected rooms into complete groups using union-find,
    so every room shows ALL rooms in its group, not just direct links."""
    all_rooms = set(travelling_together.keys())
    for rooms in travelling_together.values():
        all_rooms.update(rooms)
    parent = {r: r for r in all_rooms}

    def find(x):
        while parent[x] != x:
            parent[x] = parent[parent[x]]
            x = parent[x]
        return x

    def union(a, b):
        parent[find(a)] = find(b)

    for room, linked in travelling_together.items():
        for other in linked:
            union(room, other)

    groups = defaultdict(set)
    for room in all_rooms:
        groups[find(room)].add(room)

    room_to_group = {}
    for group in groups.values():
        if len(group) > 1:
            for room in group:
                room_to_group[room] = sorted(group)  # include self
    return room_to_group


def build_summary_page(summary_data):
    """Build a landscape A4 summary/briefing page as PDF bytes."""
    buf = io.BytesIO()
    page_size = landscape(A4)
    doc = SimpleDocTemplate(buf, pagesize=page_size,
                            leftMargin=15*mm, rightMargin=15*mm,
                            topMargin=12*mm, bottomMargin=12*mm)
    W = page_size[0] - 30*mm
    story = []

    C_HEADER  = colors.HexColor('#1a1a2e')
    C_MUT     = colors.HexColor('#6b7280')
    C_BG      = colors.HexColor('#f8fafc')
    C_BORDER  = colors.HexColor('#e2e8f0')
    C_ROW_ALT = colors.HexColor('#f1f5f9')

    def ps(size, color=C_HEADER, bold=False, align=TA_LEFT):
        return ParagraphStyle('x', fontSize=size, textColor=color,
                              fontName='Helvetica-Bold' if bold else 'Helvetica',
                              alignment=align, leading=size * 1.4)

    # Header — two rows: title line + subtitle line, well spaced
    ht = Table([[
        Paragraph(
            f'<font size="8" color="#6b7280">{summary_data["property"]}</font>',
            ps(8, C_MUT)),
        Paragraph(
            f'<font size="8" color="#6b7280">Generated: {summary_data["generated"]}  |  '
            f'{summary_data["total"]} arrivals  |  {summary_data["rooms"]} rooms</font>',
            ps(8, C_MUT, align=TA_CENTER)),
        Paragraph(
            f'<font size="8" color="#6b7280">Arrival Date: {summary_data["date"]}</font>',
            ps(8, C_MUT, align=TA_CENTER)),
    ]], colWidths=[W*0.4, W*0.35, W*0.25])
    ht.setStyle(TableStyle([
        ('BACKGROUND',(0,0),(-1,-1),C_BG),
        ('BOX',(0,0),(-1,-1),0.5,C_BORDER),
        ('TOPPADDING',(0,0),(-1,-1),6),('BOTTOMPADDING',(0,0),(-1,-1),6),
        ('LEFTPADDING',(0,0),(-1,-1),14),('RIGHTPADDING',(0,0),(-1,-1),14),
        ('VALIGN',(0,0),(-1,-1),'MIDDLE'),
    ]))

    title_row = Table([[
        Paragraph(
            '<font size="18"><b>Daily Operations Meeting</b></font>',
            ps(18, C_HEADER, bold=True))
    ]], colWidths=[W])
    title_row.setStyle(TableStyle([
        ('BACKGROUND',(0,0),(-1,-1),C_BG),
        ('LINEBELOW',(0,0),(-1,-1),0.5,C_BORDER),
        ('LINEBEFORE',(0,0),(0,-1),0.5,C_BORDER),
        ('LINEAFTER',(-1,0),(-1,-1),0.5,C_BORDER),
        ('TOPPADDING',(0,0),(-1,-1),8),('BOTTOMPADDING',(0,0),(-1,-1),8),
        ('LEFTPADDING',(0,0),(-1,-1),14),('RIGHTPADDING',(0,0),(-1,-1),14),
        ('VALIGN',(0,0),(-1,-1),'MIDDLE'),
    ]))

    story.append(title_row)
    story.append(ht)
    story.append(Spacer(1, 5))

    # Stats — match HTML stat cards
    stat_items = [
        ('Rooms',      summary_data['rooms']),
        ('Repeaters',  summary_data['counts'].get('repeater', 0)),
        ('Flights',    summary_data['counts'].get('flight', 0)),
        ('GHA',        summary_data['counts'].get('membership', 0)),
        ('VIP',        summary_data['counts'].get('vip', 0)),
        ('Complaints', summary_data['counts'].get('complaint', 0)),
        ('Allergies',  summary_data['counts'].get('allergy', 0)),
        ('Occasions',  summary_data['counts'].get('occasion', 0)),
        ('Payment',    summary_data['counts'].get('payment', 0)),
        ('Early CI',   len([g for g in summary_data['guests'] if any('Early Check-in' in f for f in g['flags'])])),
        ('Late CO',    len([g for g in summary_data['guests'] if any('Late Check-out' in f for f in g['flags'])])),
        ('Upgrade',    len([g for g in summary_data['guests'] if any('Upgrade' in f for f in g['flags'])])),
        ('Comp',       len([g for g in summary_data['guests'] if any('Comp' in f for f in g['flags'])])),
        ('Children',   len([g for g in summary_data['guests'] if any('Child' in f for f in g['flags'])])),
        ('No Flight',  len([g for g in summary_data['guests'] if g.get('flight') == 'NO FLIGHT INFO'])),
        ('Long Stay',  len([g for g in summary_data['guests'] if any('Long Stay' in f for f in g['flags'])])),
        ('Pot. VIP',   len([g for g in summary_data['guests'] if any('Potential VIP' in f for f in g['flags'])])),
    ]
    # Split into two rows of stat cards if many items
    cw = W / min(len(stat_items), 9)
    row1 = stat_items[:9]
    row2 = stat_items[9:]
    def make_stat_table(items):
        _cw = W / len(items)
        _labels  = [Paragraph(f'<font size="7" color="#6b7280">{s[0]}</font>',
                               ps(7, align=TA_CENTER)) for s in items]
        _numbers = [Paragraph(f'<b>{s[1]}</b>',
                               ps(18, C_HEADER, bold=True, align=TA_CENTER)) for s in items]
        _t = Table([_labels, _numbers], colWidths=[_cw]*len(items))
        _t.setStyle(TableStyle([
            ('BACKGROUND',(0,0),(-1,-1),C_BG),
            ('BOX',(0,0),(-1,-1),0.5,C_BORDER),
            ('LINEAFTER',(0,0),(-2,-1),0.5,C_BORDER),
            ('ALIGN',(0,0),(-1,-1),'CENTER'),
            ('VALIGN',(0,0),(-1,-1),'MIDDLE'),
            ('TOPPADDING',(0,0),(-1,0),5),('BOTTOMPADDING',(0,0),(-1,0),1),
            ('TOPPADDING',(0,1),(-1,1),1),('BOTTOMPADDING',(0,1),(-1,1),6),
        ]))
        return _t
    story.append(make_stat_table(row1))
    if row2:
        story.append(Spacer(1, 2))
        story.append(make_stat_table(row2))
    story.append(Spacer(1, 5))

    # Guest table
    def flag_para(flags):
        parts = []
        for f in flags:
            fl = f.lower()
            if any(x in f for x in ['VIP','Titanium','Platinum','Gold','Silver','Red']):
                bg, fg = '#EEEDFE','#26215C'
            elif any(x in fl for x in ['allerg','shellfish','gluten','peanut','lactose','vegan','halal']):
                bg, fg = '#FCEBEB','#501313'
            elif 'complaint' in fl or 'glitch' in fl:
                bg, fg = '#FCEBEB','#501313'
            elif 'collect' in fl or 'payment' in fl:
                bg, fg = '#FAEEDA','#633806'
            elif 'together' in fl:
                bg, fg = '#E6F1FB','#042C53'
            elif any(x in fl for x in ['rpt','repeat']):
                bg, fg = '#E1F5EE','#04342C'
            elif any(x in f for x in ['Honeymoon','Anniversary','Birthday','Wedding','Babymoon','Proposal']):
                bg, fg = '#EEEDFE','#26215C'
            elif 'leg' in fl:
                bg, fg = '#FAEEDA','#633806'
            elif 'early check-in' in fl:
                bg, fg = '#E1F5EE','#04342C'
            elif 'late check-out' in fl:
                bg, fg = '#FAEEDA','#633806'
            elif 'upgrade' in fl:
                bg, fg = '#E6F1FB','#042C53'
            elif fl == 'comp':
                bg, fg = '#FBEAF0','#72243E'
            elif 'child' in fl:
                bg, fg = '#FAEEDA','#633806'
            elif 'long stay' in fl:
                bg, fg = '#E1F5EE','#04342C'
            elif 'no flight info' in fl:
                bg, fg = '#FCEBEB','#501313'
            elif 'potential vip' in fl:
                bg, fg = '#FAEEDA','#633806'
            elif 'restaurant' in fl:
                bg, fg = '#E1F5EE','#04342C'
            elif 'spa' in fl:
                bg, fg = '#FBEAF0','#72243E'
            elif 'activit' in fl:
                bg, fg = '#E6F1FB','#042C53'
            elif 'sharer missing' in fl:
                bg, fg = '#FCEBEB','#501313'
            elif fl.startswith('poa'):
                bg, fg = '#FAEEDA','#633806'
            elif 'potential lqa' in fl:
                bg, fg = '#FCEBEB','#501313'
            elif 'payment not received' in fl:
                bg, fg = '#FCEBEB','#501313'
            else:
                bg, fg = '#F1EFE8','#2C2C2A'
            parts.append(f'<font size="7" color="{fg}" backColor="{bg}"> {f} </font>')
        return Paragraph('  '.join(parts), ps(7))

    def flight_para(f):
        if f == 'NO FLIGHT INFO':
            return Paragraph('<font size="7" color="#ffffff" backColor="#E24B4A"> ❓ No Flight Info </font>', ps(7))
        elif 'NO ETA' in f:
            return Paragraph(f'<font size="7" color="#ffffff" backColor="#E24B4A"> {f} </font>', ps(7))
        elif f and f not in ('--', '-', ''):
            return Paragraph(f'<font size="7" color="#633806" backColor="#FAEEDA"> {f} </font>', ps(7))
        return Paragraph('-', ps(7, C_MUT))

    def guest_para(g):
        name     = g['name']
        ta       = g.get('ta', '')
        checkin  = g.get('checkin', '')
        checkout = g.get('checkout', '')
        nights   = g.get('nights', 0)
        adults   = g.get('adults', 0)
        children = g.get('children', 0)
        lines = [f'<b><font size="9">{name}</font></b>']
        if ta:
            lines.append(f'<font size="7" color="#6b7280"><i>{ta}</i></font>')
        if checkin and checkout:
            night_str = f' ({nights}N)' if nights else ''
            lines.append(f'<font size="7" color="#1e2535">{checkin} → {checkout}{night_str}</font>')
        pax = f'Adl: {adults}'
        if children:
            pax += f'  Chl: {children}'
        lines.append(f'<font size="7" color="#374151">{pax}</font>')
        return Paragraph('<br/>'.join(lines), ps(8))

    def conf_para(g):
        conf      = g.get('conf', '')
        room_type = g.get('room_type', '')
        rate_code = g.get('rate_code', '')
        lines = [f'<font size="8" color="#6b7280">{conf}</font>']
        if room_type:
            lines.append(f'<font size="7" color="#374151"><b>{room_type}</b></font>')
        if rate_code:
            lines.append(f'<font size="7" color="#6b7280">{rate_code}</font>')
        return Paragraph('<br/>'.join(lines), ps(8))

    cws = [12*mm, 48*mm, 22*mm, 26*mm, 68*mm, W-176*mm]
    thead = [
        Paragraph('Room',     ps(8, C_MUT, bold=True)),
        Paragraph('Guest',    ps(8, C_MUT, bold=True)),
        Paragraph('Conf / Type', ps(8, C_MUT, bold=True)),
        Paragraph('Flight',   ps(8, C_MUT, bold=True)),
        Paragraph('Flags',    ps(8, C_MUT, bold=True)),
        Paragraph('Notes',    ps(8, C_MUT, bold=True)),
    ]
    rows = [thead]
    for g in summary_data['guests']:
        rows.append([
            Paragraph(f'<b><font size="11">{g["room"]}</font></b>', ps(9, C_HEADER, bold=True)),
            guest_para(g),
            conf_para(g),
            flight_para(g['flight']),
            flag_para(g['flags']),
            Paragraph(f'<font size="7" color="#374151">{g["note"]}</font>', ps(7)),
        ])

    gt = Table(rows, colWidths=cws, repeatRows=1)
    rstyles = [
        ('BACKGROUND',(0,0),(-1,0),C_BG),
        ('BOX',(0,0),(-1,-1),0.5,C_BORDER),
        ('INNERGRID',(0,0),(-1,-1),0.5,C_BORDER),
        ('TOPPADDING',(0,0),(-1,-1),7),('BOTTOMPADDING',(0,0),(-1,-1),7),
        ('LEFTPADDING',(0,0),(-1,-1),7),('RIGHTPADDING',(0,0),(-1,-1),7),
        ('VALIGN',(0,0),(-1,-1),'TOP'),
    ]
    for i in range(1, len(rows)):
        if i % 2 == 0:
            rstyles.append(('BACKGROUND',(0,i),(-1,i),C_ROW_ALT))
    gt.setStyle(TableStyle(rstyles))
    story.append(gt)
    story.append(Spacer(1, 5))

    # Legend
    legend_items = [
        ('#EEEDFE','#26215C','VIP / membership'),
        ('#FCEBEB','#501313','Complaint / allergy'),
        ('#FAEEDA','#633806','Payment to collect'),
        ('#E6F1FB','#042C53','Travelling together'),
        ('#E1F5EE','#04342C','Repeater / stay'),
        ('#E24B4A','#ffffff','Flight - no ETA'),
    ]
    lparts = ['<font size="8" color="#6b7280"><b>Legend:  </b></font>']
    for bg, fg, label in legend_items:
        lparts.append(f'<font size="8" color="{fg}" backColor="{bg}"> {label} </font>  ')
    lt = Table([[Paragraph(''.join(lparts), ps(8))]], colWidths=[W])
    lt.setStyle(TableStyle([
        ('BACKGROUND',(0,0),(-1,-1),C_BG),
        ('BOX',(0,0),(-1,-1),0.5,C_BORDER),
        ('TOPPADDING',(0,0),(-1,-1),7),('BOTTOMPADDING',(0,0),(-1,-1),7),
        ('LEFTPADDING',(0,0),(-1,-1),10),
    ]))
    story.append(lt)

    doc.build(story)
    buf.seek(0)
    return buf.getvalue()


def highlight_pdf(pdf_bytes, enabled_cats):
    reader = PdfReader(io.BytesIO(pdf_bytes))
    writer = PdfWriter()
    writer.append(reader)
    counts = {k: 0 for k in CATEGORIES}
    total  = 0

    with pdfplumber.open(io.BytesIO(pdf_bytes)) as pdf:

        # ══ PASS 1: identify all bookings & sharers per page ═════════════
        # booking = {room, conf, pg_idx, row_top, row_bottom, is_sharer}
        all_bookings = []

        for pg_idx, page in enumerate(pdf.pages):
            ph = float(page.height)
            words = [w for w in page.extract_words(x_tolerance=2, y_tolerance=2)
                     if w['top'] < ph * 0.91]
            processed = set()
            for rw in [w for w in words if w['x0'] < 40
                       and re.match(r'^[0-9]{1,4}$', w['text'])]:
                t  = rw['top']
                rk = round(t)
                if rk in processed: continue
                processed.add(rk)
                row = sorted([w for w in words if abs(w['top'] - t) <= 4],
                             key=lambda x: x['x0'])
                adl  = next((w for w in row if 415 <= w['x0'] <= 435
                             and w['text'].isdigit()), None)
                name = next((w for w in row if 41 <= w['x0'] <= 280), None)
                conf = next((w for w in words if w['top'] > t+3 and w['top'] < t+25
                             and 35 <= w['x0'] <= 180
                             and re.match(r'^[0-9]{6,}$', w['text'])), None)
                adl_val   = int(adl['text']) if adl else -1
                has_star  = bool(name and name['text'].startswith('*'))
                is_sharer = has_star and adl_val == 0
                all_bookings.append({
                    'room':      rw['text'],
                    'conf':      conf['text'] if conf else '',
                    'pg_idx':    pg_idx,
                    'row_top':   t,
                    'is_sharer': is_sharer,
                })

        # ── Build sharer y-ranges per page ────────────────────────────────
        # For each page, collect (top_start, top_end) ranges belonging to sharers
        # A sharer's range = from its row_top to just before the next booking row
        sharer_ranges = defaultdict(list)  # pg_idx → list of (start, end)

        for pg_idx in set(b['pg_idx'] for b in all_bookings):
            page_bkgs = sorted([b for b in all_bookings if b['pg_idx'] == pg_idx],
                               key=lambda x: x['row_top'])
            for i, b in enumerate(page_bkgs):
                if not b['is_sharer']:
                    continue
                top_start = b['row_top'] - 2  # include the header row itself
                top_end   = (page_bkgs[i+1]['row_top'] - 2
                             if i+1 < len(page_bkgs) else 999)
                sharer_ranges[pg_idx].append((top_start, top_end))

        def is_sharer_line(pg_idx, line_top):
            """Return True if this line falls within a sharer booking's range."""
            for (s, e) in sharer_ranges.get(pg_idx, []):
                if s <= line_top < e:
                    return True
            return False

        # ══ PASS 2: build travelling-together map ════════════════════════
        all_rooms = {b['room'] for b in all_bookings}
        travelling_together = defaultdict(set)

        # Link by shared conf number across bookings
        conf_to_rooms = defaultdict(set)
        for b in all_bookings:
            if b['conf']:
                conf_to_rooms[b['conf']].add(b['room'])
        for conf, rooms in conf_to_rooms.items():
            if len(rooms) > 1:
                for r in rooms:
                    travelling_together[r].update(rooms - {r})

        # Link by conf/room numbers mentioned in comments
        for pg_idx, page in enumerate(pdf.pages):
            ph = float(page.height)
            words = [w for w in page.extract_words(x_tolerance=2, y_tolerance=2)
                     if w['top'] < ph * 0.91]
            page_bkgs = sorted([b for b in all_bookings if b['pg_idx'] == pg_idx],
                               key=lambda x: x['row_top'])
            for i, b in enumerate(page_bkgs):
                top_start    = b['row_top'] + 20
                top_end      = (page_bkgs[i+1]['row_top']
                                if i+1 < len(page_bkgs) else ph * 0.91)
                comment_text = ' '.join(w['text'] for w in words
                                        if top_start <= w['top'] <= top_end)
                for other_b in all_bookings:
                    if other_b['room'] == b['room'] and other_b['conf'] == b['conf']:
                        continue
                    # Only link via conf number mentions in comments — room number
                    # scanning causes false links from adjacent booking headers
                    if other_b['conf'] and re.search(
                            r'\b' + re.escape(other_b['conf']) + r'\b', comment_text):
                        travelling_together[b['room']].add(other_b['room'])
                        travelling_together[other_b['room']].add(b['room'])

            # Also link rooms sharing the same T# ticket number in comments
            t_refs = re.findall(r'T#\s*(\d{5,})', comment_text)
            if t_refs:
                for other_b in all_bookings:
                    if other_b['room'] == b['room'] and other_b['conf'] == b['conf']:
                        continue
                    other_top_start = other_b['row_top'] + 20
                    other_pg = pdf.pages[other_b['pg_idx']]
                    other_ph = float(other_pg.height)
                    other_words = [w for w in other_pg.extract_words(
                        x_tolerance=2, y_tolerance=2) if w['top'] < other_ph * 0.91]
                    other_page_bkgs = sorted(
                        [bk for bk in all_bookings if bk['pg_idx'] == other_b['pg_idx']],
                        key=lambda x: x['row_top'])
                    other_idx = next((j for j, bk in enumerate(other_page_bkgs)
                                      if bk['room'] == other_b['room']
                                      and bk['conf'] == other_b['conf']), None)
                    if other_idx is None: continue
                    other_top_end = (other_page_bkgs[other_idx+1]['row_top']
                                     if other_idx+1 < len(other_page_bkgs)
                                     else other_ph * 0.91)
                    other_comment = ' '.join(w['text'] for w in other_words
                                             if other_top_start <= w['top'] <= other_top_end)
                    other_t_refs = re.findall(r'T#\s*(\d{5,})', other_comment)
                    # If they share at least one T# reference, link them
                    if set(t_refs) & set(other_t_refs):
                        travelling_together[b['room']].add(other_b['room'])
                        travelling_together[other_b['room']].add(b['room'])

        # Merge all connected rooms into complete groups
        room_to_group = merge_groups(travelling_together)

        # ══ PASS 3: highlight each page ══════════════════════════════════
        for pg_idx, page in enumerate(pdf.pages):
            ph       = float(page.height)
            footer_y = ph * 0.91
            words    = [w for w in page.extract_words(x_tolerance=2, y_tolerance=2)
                        if w['top'] < footer_y]
            annots   = []

            def hl(x0, top, x1, bottom, cat, color=YELLOW, popup=None):
                nonlocal total
                counts[cat] += 1
                total += 1
                annots.append(make_annot(x0, top, x1, bottom, ph, color, popup))

            # Build lines
            lg = defaultdict(list)
            for w in words: lg[round(w['top'])].append(w)
            lines = []
            for k in sorted(lg):
                ws = sorted(lg[k], key=lambda x: x['x0'])
                lines.append({
                    'text':   ' '.join(w['text'] for w in ws),
                    'x0':     min(w['x0'] for w in ws),
                    'x1':     max(w['x1'] for w in ws),
                    'top':    min(w['top'] for w in ws),
                    'bottom': max(w['bottom'] for w in ws),
                })

            seen = set()
            def hl_line(ln, cat, color=YELLOW, popup=None):
                # Skip if this line belongs to a sharer booking
                if is_sharer_line(pg_idx, ln['top']):
                    return
                k = f"{ln['x0']:.0f}_{ln['top']:.0f}"
                if k in seen: return
                seen.add(k)
                hl(ln['x0'], ln['top'], ln['x1'], ln['bottom'], cat, color, popup)

            # Keywords
            kw_cats = {k: v['kw'] for k, v in CATEGORIES.items() if v.get('kw')}
            for cat_id, kws in kw_cats.items():
                if not enabled_cats.get(cat_id, True): continue
                for ln in lines:
                    if any(kw.lower() in ln['text'].lower() for kw in kws):
                        hl_line(ln, cat_id)

            # Highlight Specials lines that contain occasion codes
            if enabled_cats.get('occasion'):
                _OCC_CODES = {'HON','HMB','HMA','WED','BDC','BEN','BTD','ANN'}
                for ln in lines:
                    if re.match(r'Specials:', ln['text'], re.I):
                        codes = set(re.findall(r'[A-Z0-9]{2,4}', ln['text']))
                        if codes & _OCC_CODES:
                            hl_line(ln, 'occasion')

            # Auto patterns
            for ln in lines:
                t = ln['text']
                if enabled_cats.get('flight'):
                    fm = re.search(
                        r'(EK\s*\d{3,4}|EY\s*\d{3,4}|QR\s*\d{3,4}|'
                        r'G9\s*\d{3,4}|KU\s*\d{3,4}|GF\s*\d{3,4})', t, re.I)
                    if fm:
                        has_eta = bool(re.search(
                            r'\b([0-1]?[0-9]|2[0-3]):[0-5][0-9]\b', t))
                        hl_line(ln, 'flight',
                                YELLOW if has_eta else ORANGE,
                                None if has_eta else '⚠️ ETA MISSING — please check manually')
                if enabled_cats.get('repeater') and re.search(
                        r'\d+(st|nd|rd|th)\s+[Tt]ime\s+RPT|\bRPT\b', t):
                    hl_line(ln, 'repeater')
                if enabled_cats.get('stayhistory') and re.search(
                        r'\d+(st|nd|rd|th)\s+[Ss]tay|Upcoming\s+[Ss]tay', t, re.I):
                    hl_line(ln, 'stayhistory')
                if enabled_cats.get('membership') and re.search(
                        r'Membership\s+Level\s+(GOLD|PLATINUM|TITANIUM|RED)', t, re.I):
                    hl_line(ln, 'membership')
                if enabled_cats.get('dbalance') and re.search(r'[1-9]\d*D\$', t, re.I):
                    hl_line(ln, 'dbalance')
                if enabled_cats.get('vip') and re.search(r'VIP[A-Z0-9]', t):
                    hl_line(ln, 'vip')
                if enabled_cats.get('sharewith') and re.search(r'Share\s+with:', t, re.I):
                    if len(re.sub(r'Share\s+with:', '', t, flags=re.I).strip()) > 1:
                        hl_line(ln, 'sharewith')
                if enabled_cats.get('earlyarr') and re.search(r'0[0-8]:\d{2}', t) \
                        and not re.search(r'Arrival\s+Time', t, re.I):
                    hl_line(ln, 'earlyarr')
                # Potential VIP — title in notes, no VIP code on room/conf row
                if enabled_cats.get('vip'):
                    _vip_title_pat = (
                        r'\bCEO\b|\bCOO\b|\bCFO\b|\bCTO\b|\bCMO\b|\bHRH\b'
                        r'|His\s+Excellency|Her\s+Excellency|\bH\.E\.\b'
                        r'|Vice\s+President|Managing\s+Director'
                        r'|\bAmbassador\b|Prime\s+Minister|\bGovernor\b|\bSenator\b'
                    )
                    if re.search(_vip_title_pat, t, re.I):
                        hl_line(ln, 'vip', ORANGE, '⚑ Potential VIP — title mentioned in notes')
                # Restaurant booking mentions — specific names only
                # Exclude Fixed Charges, Adj., number+Adj., SPO, discount/rate lines
                _excl_pat = (r'Fixed\s+Charges|^\d+\s+Adj\.|^Adj\.|SPO:'
                             r'|CONTRACTED|DISCOUNT\s+APPLIED|Discount\s+Offer'
                             r'|Offered\s+Applied|F&B\s+Discount|Spa\s+Discount')
                _rest_hl_pats = (
                    r'Fushi\s+Caf[eé]|Sea\s+Fire\s+Salt|\bSFS\b|\bAqua\b'
                    r'|Yellow\s+Fin\s+Club|\bOrigami\b|Dhoni\s+Bar|\bCumin\b'
                    r'|Baan\s+Huraa|Designer\s+Dining|Living\s+Room'
                )
                if (re.search(_rest_hl_pats, t, re.I)
                        and not re.match(_excl_pat, t.strip(), re.I)
                        and not re.search(r'complaint|glitch|upset|charged.*restaurant', t, re.I)):
                    hl_line(ln, 'occasion', YELLOW, '🍽️ Restaurant booking mentioned')
                # Spa booking mentions — explicit booking language only
                _spa_hl_pats = (
                    r'Anantara\s+Spa|Dhigu\s+Spa|Veli\s+Spa'
                    r'|spa\s+(?:booking|reserv|treatment|session)'
                    r'|(?:book|reserv)\w*\s+(?:at\s+)?(?:the\s+)?spa\b'
                    r'|(?:couples?|four\s+hands?|hot\s+stone|body\s+scrub|facial|swedish|deep\s+tissue)\s+massage'
                )
                if (re.search(_spa_hl_pats, t, re.I)
                        and not re.match(_excl_pat, t.strip(), re.I)):
                    hl_line(ln, 'occasion', YELLOW, '💆 Spa booking mentioned')
                # Activity mentions
                _act_hl_pats = (
                    r'snorkel|scuba|\bdiv(?:ing?|e)\b|sunset\s+(?:cruise|sail|dhoni)'
                    r'|fishing\s+trip|\bdolphin\b|\bwhale\b|\bkayak'
                    r'|watersport|water\s+sport|sandbank|overwater\s+dinner'
                    r'|Activities\s+-\s+(?:Water|Diving|Snorkel)'
                )
                if (re.search(_act_hl_pats, t, re.I)
                        and not re.match(_excl_pat, t.strip(), re.I)):
                    hl_line(ln, 'occasion', YELLOW, '🤿 Activity mentioned')

            # Room / Name / Conf No — skip sharers entirely
            if enabled_cats.get('roomno'):
                processed = set()
                for rw in [w for w in words if w['x0'] < 40
                           and re.match(r'^[0-9]{1,4}$', w['text'])]:
                    t  = rw['top']
                    rk = round(t)
                    if rk in processed: continue
                    # Skip sharer rows
                    if is_sharer_line(pg_idx, t):
                        processed.add(rk)
                        continue
                    row = sorted([w for w in words if abs(w['top'] - t) <= 4],
                                 key=lambda x: x['x0'])
                    adl = next((w for w in row if 415 <= w['x0'] <= 435
                                and w['text'].isdigit()), None)
                    if not adl or int(adl['text']) == 0:
                        continue
                    name = [w for w in row if 41 <= w['x0'] <= 280]
                    if not name: continue
                    processed.add(rk)

                    together = room_to_group.get(rw['text'], [])
                    popup = ('👥 Travelling Together with Room(s): ' +
                             ', '.join(together)) if together else None

                    hl(rw['x0'], rw['top'], rw['x1'], rw['bottom'], 'roomno', YELLOW, popup)
                    hl(min(w['x0'] for w in name), min(w['top'] for w in name),
                       max(w['x1'] for w in name), max(w['bottom'] for w in name),
                       'roomno', YELLOW, popup)
                    conf = next((w for w in words if w['top'] > t+3 and w['top'] < t+25
                                 and 35 <= w['x0'] <= 180
                                 and re.match(r'^[0-9]{6,}$', w['text'])), None)
                    if conf:
                        hl(conf['x0'], conf['top'], conf['x1'], conf['bottom'],
                           'roomno', YELLOW, popup)

            if annots:
                p = writer.pages[pg_idx]
                if '/Annots' in p:
                    for a in annots: p['/Annots'].append(a)
                else:
                    p[NameObject('/Annots')] = ArrayObject(annots)

        # ══ PASS 4: collect summary guest data ═══════════════════════════
        # Get property name and date from first page
        first_page_words = pdf.pages[0].extract_words(x_tolerance=2, y_tolerance=2)
        property_name = 'Arrival Report'
        arrival_date  = date.today().strftime('%d %B %Y')
        generated     = ''
        for w in first_page_words:
            if 'Anantara' in w['text'] or 'Naladhu' in w['text']:
                # Get full line
                t = w['top']
                line_words = sorted([x for x in first_page_words if abs(x['top']-t)<=3], key=lambda x: x['x0'])
                property_name = ' '.join(x['text'] for x in line_words
                                         if not re.match(r'\d{2}/\d{2}/\d{2}', x['text']))
                break
        # Find arrival date from report
        for w in first_page_words:
            if re.match(r'\d{2}/\d{2}/\d{2}', w['text']):
                generated = w['text']
                break
        for w in first_page_words:
            if w['text'] == 'Arrival' and w['x0'] < 100:
                # next token should be 'Date' then the date
                t = w['top']
                line = sorted([x for x in first_page_words if abs(x['top']-t)<=3], key=lambda x: x['x0'])
                line_text = ' '.join(x['text'] for x in line)
                dm = re.search(r'(\d{2}/\d{2}/\d{2})', line_text)
                if dm:
                    parts = dm.group(1).split('/')
                    months = ['','Jan','Feb','Mar','Apr','May','Jun','Jul','Aug','Sep','Oct','Nov','Dec']
                    try:
                        arrival_date = f"{int(parts[0])} {months[int(parts[1])]} 20{parts[2]}"
                    except: pass
                break

        # Collect one row per non-sharer booking with adults > 0
        summary_guests = []
        for pg_idx, page in enumerate(pdf.pages):
            ph = float(page.height)
            words = [w for w in page.extract_words(x_tolerance=2, y_tolerance=2)
                     if w['top'] < ph * 0.91]
            page_bkgs = sorted([b for b in all_bookings if b['pg_idx'] == pg_idx],
                               key=lambda x: x['row_top'])

            processed_s = set()
            for rw in [w for w in words if w['x0'] < 40
                       and re.match(r'^[0-9]{1,4}$', w['text'])]:
                t  = rw['top']
                rk = round(t)
                if rk in processed_s: continue
                processed_s.add(rk)

                row = sorted([w for w in words if abs(w['top'] - t) <= 4],
                             key=lambda x: x['x0'])
                adl  = next((w for w in row if 415 <= w['x0'] <= 435
                             and w['text'].isdigit()), None)
                name = next((w for w in row if 41 <= w['x0'] <= 280), None)
                if not adl or int(adl['text']) == 0: continue
                if not name: continue
                if is_sharer_line(pg_idx, t): continue

                # Travel Agent — on the line 1-3pts above room row, x0 ~182-270
                ta_line = [w for w in words if abs(w['top'] - (t - 1)) <= 4
                           and 182 <= w['x0'] <= 280]
                ta_name = ' '.join(w['text'] for w in sorted(ta_line, key=lambda x: x['x0']))
                # Strip the source prefix (T-, C-, W-, etc)
                ta_name = re.sub(r'^[A-Z]-\s*', '', ta_name).strip()

                # Conf
                conf = next((w for w in words if w['top'] > t+3 and w['top'] < t+25
                             and 35 <= w['x0'] <= 180
                             and re.match(r'^[0-9]{6,}$', w['text'])), None)

                # Flight + ETA + arrival method on conf sub-row
                conf_line_words = sorted([w for w in words if w['top'] > t+3 and w['top'] < t+25],
                                         key=lambda x: x['x0'])

                # Flight: standard combined (GF144, 6E1133) or split (EK + 656, AI + 239)
                flight_text = ''
                flight_top  = None
                flight_w = next((w for w in conf_line_words
                                 if 280 < w['x0'] < 380
                                 and re.match(r'^[A-Z0-9]{2,3}\d{2,4}$', w['text'])), None)
                if flight_w:
                    flight_text = flight_w['text']
                    flight_top  = flight_w['top']
                else:
                    # Split: two-letter alpha code immediately followed by digit-only word
                    code_w = next((w for w in conf_line_words if 280 < w['x0'] < 360
                                   and re.match(r'^[A-Z]{2,3}$', w['text'])), None)
                    if code_w:
                        num_w = next((w for w in conf_line_words
                                      if abs(w['top'] - code_w['top']) <= 3
                                      and code_w['x0'] < w['x0'] < code_w['x0'] + 35
                                      and re.match(r'^\d{2,4}$', w['text'])), None)
                        if num_w:
                            flight_text = f"{code_w['text']}{num_w['text']}"
                            flight_top  = code_w['top']

                # Non-flight transport codes at flight position
                if not flight_text:
                    trans_w = next((w for w in conf_line_words
                                    if 280 < w['x0'] < 380
                                    and re.match(r'^(RMV|OTH|Airport|FERRY|TRF)$', w['text'], re.I)), None)
                    if trans_w:
                        flight_text = trans_w['text']
                        flight_top  = trans_w['top']

                flight_str = ''
                if flight_text:
                    eta_w = next((w for w in conf_line_words
                                  if flight_top and abs(w['top'] - flight_top) <= 3
                                  and re.match(r'^([01]?\d|2[0-3]):[0-5]\d$', w['text'])), None)
                    if eta_w:
                        flight_str = f"{flight_text} {eta_w['text']}"
                    elif re.match(r'^(RMV|OTH|Airport|FERRY|TRF)$', flight_text, re.I):
                        flight_str = flight_text  # no ETA concept for these
                    else:
                        flight_str = f"{flight_text} NO ETA"
                else:
                    flight_str = 'NO FLIGHT INFO'

                # Arrival method: SBA/SBR/RMV/OTH/etc at x≈375-430
                arr_method_w = next((w for w in conf_line_words
                                     if 375 <= w['x0'] <= 430
                                     and re.match(r'^[A-Z]{2,4}$', w['text'])
                                     and w['text'] not in ('VIP',)), None)
                arr_method = arr_method_w['text'] if arr_method_w else ''

                # Collect flags for this booking
                bflags = []
                # Check categories from counts — scan comment text
                i_bkg = next((j for j, b in enumerate(page_bkgs)
                              if round(b['row_top']) == rk), None)
                top_start = t + 20
                top_end   = (page_bkgs[i_bkg+1]['row_top']
                             if i_bkg is not None and i_bkg+1 < len(page_bkgs)
                             else ph * 0.91)
                ct = ' '.join(w['text'] for w in words if top_start <= w['top'] <= top_end)

                # Full name + TA text for title/allergy scanning
                name_text = ' '.join(w['text'] for w in row if 41 <= w['x0'] <= 280)
                ta_full   = ta_name  # already extracted above

                if re.search(r'Membership\s+Level\s+(GOLD|PLATINUM|TITANIUM|RED|SILVER)', ct, re.I):
                    m = re.search(r'Membership\s+Level\s+(\w+)', ct, re.I)
                    if m: bflags.append(m.group(1).capitalize())
                # VIP: check room row AND conf sub-row (y+10 to y+20)
                vip_search_text = ' '.join(w['text'] for w in words if abs(w['top']-t)<=20 and w['x0']>30)
                if re.search(r'VIP[A-Z0-9]', vip_search_text):
                    vip_w = next((w for w in words if abs(w['top']-t)<=20 and re.match(r'VIP[A-Z0-9]', w['text'])), None)
                    if vip_w: bflags.append(vip_w['text'])
                if re.search(r'\d+(st|nd|rd|th)\s+[Tt]ime\s+RPT|\bRPT\b', ct):
                    m = re.search(r'(\d+(?:st|nd|rd|th)\s+[Tt]ime\s+RPT|\bRPT\b)', ct)
                    if m: bflags.append(m.group(1))

                # ── Improvement 6: Repeat guest from stay history ────────
                _stays_m = re.search(r'Total stays across chain\s*(\d+)', ct)
                if _stays_m and int(_stays_m.group(1)) > 1:
                    if not any('RPT' in f or 'Repeat' in f for f in bflags):
                        bflags.append(f"RPT ({_stays_m.group(1)} stays)")

                # ── Occasions: text scan ─────────────────────────────────
                for kw in ['birthday','anniversary','honeymoon','babymoon','wedding','proposal','engagement']:
                    if kw in ct.lower(): bflags.append(kw.capitalize()); break

                # ── Improvement 5: Special occasion codes from Specials ──
                specials_m = re.search(r'Specials:\s*([A-Z0-9,]+)', ct)
                specials_codes = specials_m.group(1).split(',') if specials_m else []
                _OCCASION_CODES = {
                    'HON':'Honeymoon', 'HMB':'Honeymoon Setup',
                    'HMA':'Honeymoon Amenity', 'WED':'Wedding',
                    'BDC':'Birthday Cake', 'BEN':'Birthday Amenity',
                    'BTD':'Birthday Decoration', 'ANN':'Anniversary',
                }
                for _code, _label in _OCCASION_CODES.items():
                    if _code in specials_codes:
                        if not any(_label.split()[0].lower() in f.lower() for f in bflags):
                            bflags.append(_label)
                        break

                # ── Allergy: notes text ──────────────────────────────────
                for kw in ['allerg','shellfish','gluten','peanut','lactose','vegan','vegetarian','halal']:
                    if kw in ct.lower():
                        m = re.search(r'(\w*' + kw + r'\w*)', ct, re.I)
                        bflags.append(m.group(1).capitalize() if m else 'Allergy'); break

                # ── Improvement 4: Allergy keyword in name field ─────────
                if not any('llerg' in f.lower() or any(x in f.lower() for x in
                           ['shellfish','gluten','peanut','lactose','vegan','halal'])
                           for f in bflags):
                    for kw in ['allerg','shellfish','gluten','peanut','lactose','vegan','halal','nut','dairy']:
                        if kw in name_text.lower():
                            bflags.append(f'Allergy (name field)')
                            break

                if re.search(r'complaint|glitch|feedback|upset', ct, re.I):
                    bflags.append('Complaint')
                if re.search(r'collect upon arrival|to be collected|pls collect|please collect', ct, re.I):
                    bflags.append('Collect')
                if re.search(r'\d+(st|nd|rd|th)\s+[Ll]eg|1ST\s+Leg|2nd\s+Leg', ct):
                    bflags.append('Multi-leg')
                # D$ balance
                d_match = re.search(r'D\$\s*([0-9]+(?:\.[0-9]+)?)', ct)
                if d_match and float(d_match.group(1)) > 0:
                    bflags.append(f"D$ {d_match.group(1)}")
                # No flight info — only flag when Carr. Code is blank or SBA/SBR
                # Airport/APT/OTH/RMV/FERRY means seaplane/other transport — no flight expected
                _carr_is_air_transport = bool(re.search(
                    r'\b(Airport|APT|OTH|RMV|FERRY|TRF)\b', arr_method, re.I))
                if not flight_text and not _carr_is_air_transport:
                    bflags.append('No Flight Info')

                # ── Restaurant Bookings ──────────────────────────────────
                _REST_NAMES = [
                    r'Fushi\s+Caf[eé]', r'Sea\s+Fire\s+Salt', r'\bSFS\b',
                    r'\bAqua\b', r'Yellow\s+Fin\s+Club', r'\bOrigami\b',
                    r'Dhoni\s+Bar', r'\bCumin\b', r'Baan\s+Huraa',
                    r'Designer\s+Dining', r'Living\s+Room',
                ]

                def _is_excluded_line(ln):
                    """Lines to skip for restaurant/spa/activity scanning."""
                    return bool(re.match(
                        r'Fixed\s+Charges|Adj\.|^\d+\s+Adj\.'
                        r'|SPO:|CONTRACTED\s+RATES|DISCOUNT\s+APPLIED'
                        r'|Discount\s+Offer|Offered\s+Applied'
                        r'|F&B\s+Discount|Spa\s+Discount'
                        r'|Board\s+description|Contract\s+(?:Name|Description)'
                        r'|Routing\s+Instruction|Membership\s+Type',
                        ln.strip(), re.I))

                # Build filtered line groups once for all three checks
                _note_line_groups = defaultdict(list)
                for _w in words:
                    if top_start <= _w['top'] <= top_end:
                        _note_line_groups[round(_w['top'])].append(_w)

                _clean_lines = []
                for _tk in sorted(_note_line_groups):
                    _ln = ' '.join(_w['text'] for _w in
                                   sorted(_note_line_groups[_tk], key=lambda x: x['x0']))
                    if not _is_excluded_line(_ln):
                        _clean_lines.append(_ln)

                if not any('Restaurant' in f for f in bflags):
                    for _ln in _clean_lines:
                        if re.search(r'complaint|glitch|upset|charged.*restaurant|restaurant.*charge', _ln, re.I):
                            continue
                        for _pat in _REST_NAMES:
                            if re.search(_pat, _ln, re.I):
                                bflags.append('Restaurant Booking')
                                break
                        if any('Restaurant' in f for f in bflags): break

                # ── Spa Bookings ─────────────────────────────────────────
                # Require explicit booking language — not just "spa" as a category word
                _SPA_PATS = [
                    r'Anantara\s+Spa', r'Dhigu\s+Spa', r'Veli\s+Spa',
                    r'spa\s+(?:booking|reserv|treatment|appoint|session)',
                    r'(?:book|reserv)\w*\s+(?:at\s+)?(?:the\s+)?spa\b',
                    r'(?:couples?|four\s+hands?|hot\s+stone|body\s+scrub|facial|swedish|deep\s+tissue)\s+massage',
                    r'spa\s+(?:package|voucher|credit)',
                ]
                if not any('Spa' in f for f in bflags):
                    for _ln in _clean_lines:
                        for _pat in _SPA_PATS:
                            if re.search(_pat, _ln, re.I):
                                bflags.append('Spa Booking')
                                break
                        if any('Spa' in f for f in bflags): break

                # ── Activities ───────────────────────────────────────────
                _ACT_PATS = [
                    r'snorkel', r'scuba', r'\bdiv(?:ing?|e)\b',
                    r'sunset\s+(?:cruise|sail|dhoni|trip|fishing)',
                    r'sunrise\s+(?:cruise|fishing|trip)',
                    r'fishing\s+trip', r'\bdolphin\b', r'\bwhale\b',
                    r'\bkayak', r'\bsurf(?:ing)?\b',
                    r'water\s*sport', r'watersport',
                    r'sandbank\s+(?:picnic|trip|lunch|dinner)',
                    r'beach\s+(?:dinner|bbq|picnic)',
                    r'overwater\s+dinner', r'private\s+(?:dinner|picnic)',
                    r'Activities\s+-\s+(?:Water|Diving|Snorkel|Sport)',
                ]
                if not any('Activit' in f for f in bflags):
                    _act_ct = ' '.join(_clean_lines)
                    _acts_found = []
                    for _pat in _ACT_PATS:
                        _m = re.search(_pat, _act_ct, re.I)
                        if _m:
                            _acts_found.append(_m.group().strip())
                    if _acts_found:
                        bflags.append(f'Activities ({_acts_found[0]})')

                # ── Potential VIP — improved: check notes, name, TA ─────
                _VIP_ALREADY = any('VIP' in f or f in ('Platinum','Titanium','Gold','Silver','Red')
                                   for f in bflags)
                if not _VIP_ALREADY:
                    _VIP_TITLE_PATS = [
                        r'\bCEO\b', r'\bCOO\b', r'\bCFO\b', r'\bCTO\b', r'\bCMO\b',
                        r'\bHRH\b', r'His\s+Excellency', r'Her\s+Excellency', r'\bH\.E\.\b',
                        r'Vice\s+President', r'Vice\s+Pres\b',
                        r'Managing\s+Director',
                        r'\bAmbassador\b', r'Prime\s+Minister',
                        r'\bGovernor\b', r'\bSenator\b',
                        # ── Improvement 2: TA/company signals ────────────
                        r'Royal\s+Household', r'Embassy\s+of', r'Ministry\s+of',
                        r"Prime\s+Minister'?s?\s+Office",
                        r'State\s+Department', r'Presidential',
                        r'Minister\s+of',r'Deputy\s+Minister',r'State\s+Minister',
                    ]
                    # ── Improvement 1: scan name field + TA field too ────
                    _scan_texts = [
                        ('notes', ct),
                        ('name',  name_text),
                        ('TA',    ta_full),
                    ]
                    for _src, _txt in _scan_texts:
                        for _pat in _VIP_TITLE_PATS:
                            _m = re.search(_pat, _txt, re.I)
                            if _m:
                                _label = _m.group().strip()
                                bflags.append(f'Potential VIP ({_label})')
                                break
                        else:
                            continue
                        break

                # Early check-in / Late check-out
                if 'ECI' in specials_codes or re.search(r'\bECI\b', ct):
                    bflags.append('Early Check-in')
                if 'LCO' in specials_codes or re.search(r'\bLCO\b', ct):
                    bflags.append('Late Check-out')
                # Upgrade
                if 'UPG' in specials_codes or re.search(r'\bupgrade\b', ct, re.I):
                    bflags.append('Upgrade')
                # ── Improvement 3: COMP — distinguish package vs standalone
                _comp_standalone = bool(re.search(r'(?<![A-Z/])\bCOMP\b(?![/A-Z])', ct))
                _comp_specials   = 'COMP' in specials_codes
                _comp_phrase     = bool(re.search(r'complimentary\s+(?:stay|room|night|upgrade)', ct, re.I))
                if _comp_specials or _comp_standalone or _comp_phrase:
                    bflags.append('Comp')
                # Children
                chl_w = next((w for w in row if 437 <= w['x0'] <= 460 and w['text'].isdigit()), None)
                if chl_w and int(chl_w['text']) > 0:
                    bflags.append(f"Child ×{chl_w['text']}")
                together = room_to_group.get(rw['text'], [])
                if together:
                    bflags.append(f"Together: {', '.join(together)}")

                # ── Payment Not Received ─────────────────────────────────
                # Credit TAs — bookings from these don't need deposit upfront
                _CREDIT_TAS = {
                    'avc club developer', 'abercrombie & kent', 'abercrombie and kent',
                    'british airways holidays', 'catai tours', 'classic vacations',
                    'der touristik', 'descubre viajes', 'destinology',
                    'dnata', 'dnata holidays', 'dnata travel',
                    'helloworld', 'helloword travel',
                    'hotel beds', 'hotelbeds',
                    'hummingbird', 'hummingbird heymaa',
                    'intour maldives', 'inpac japan',
                    'international rawnaq', 'rawnaq travel',
                    'silverjet', 'silverjet vakanties',
                    'trailfinders', 'travel counsellors',
                    'east europe travel', 'tui deutschland', 'tui suisse',
                    'voyage prive', 'voyage privé',
                    # Additional exclusions
                    'traveltino',
                    'expedia',
                    'kognitiv',
                    'tbo holidays',
                    'tbo',
                    'agoda',
                    'agoda services',
                    'dertour',
                    'der tour',
                }
                # Extract deposit value at x0≈741-760
                deposit_w = next((w for w in row if 735 <= w['x0'] <= 780), None)
                deposit_val = 0.0
                if deposit_w:
                    try:
                        deposit_val = float(re.sub(r'[^\d.]', '', deposit_w['text']))
                    except Exception:
                        deposit_val = 0.0

                # Check if TA is on credit list — used later after src_code is extracted
                _ta_lower = ta_name.lower().strip()
                _on_credit = any(ct_name in _ta_lower or _ta_lower in ct_name
                                 for ct_name in _CREDIT_TAS)
                # A booking is "sharer missing" if NO adl=0 row exists for
                # this room number anywhere in the entire PDF (all_bookings)
                _room_has_sharer = any(
                    b['room'] == rw['text'] and b['is_sharer']
                    for b in all_bookings
                )
                if not _room_has_sharer:
                    bflags.append('Sharer Missing')

                # ── POA with Amount ──────────────────────────────────────
                # Only flag when a USD amount appears as part of the POA instruction
                # e.g. "POA/2GT/2TRRT (USD 779.85)" or "POA/2GT USD 500"
                # NOT "POA/2BB/2TRRT ... USD 100" (unrelated benefit line)
                _poa_m = re.search(
                    r'POA[/\w]*\s*\(USD\s*([\d,]+\.?\d*)\)'   # POA/xx (USD 779.85)
                    r'|POA[/\w]*\s+USD\s*([\d,]+\.?\d*)',       # POA/xx USD 779.85
                    ct)
                if _poa_m:
                    _amt = _poa_m.group(1) or _poa_m.group(2) or ''
                    if not any('Collect' in f for f in bflags):
                        bflags.append(f'POA: USD {_amt}' if _amt else 'POA Amount')

                # Notes — capture ALL comment lines, clean up and join
                from collections import defaultdict as _dd
                note_lg = _dd(list)
                for w in words:
                    if top_start <= w['top'] <= top_end:
                        note_lg[round(w['top'])].append(w)
                note_lines = []
                for k in sorted(note_lg.keys()):
                    ws = sorted(note_lg[k], key=lambda x: x['x0'])
                    line_text = ' '.join(w['text'] for w in ws).strip()

                    # ── Hard skip: BFB/NFB/BTR package codes ────────────
                    if re.match(r'^[\-\d\*]*[A-Z]{2,4}\d{2}[A-Z]', line_text): continue   # BFB01E, NFB02I
                    if re.match(r'^[\-\d\*]+[A-Z]{3}', line_text): continue                # -1*BFB40E
                    if re.match(r'^[A-Z]{2,4}\d{2,}[A-Z],[A-Z]', line_text): continue     # BFB40E,NFB
                    if re.match(r'^,[A-Z]{2,4}\d{2}', line_text): continue                 # ,BFB14E,BFB40 (fragment)
                    if re.match(r'^[A-Z],[A-Z]{2,4}\d{2}', line_text): continue            # E,BFB40E (fragment)
                    if re.match(r'^[A-Z]{1,2},', line_text) and re.search(r'[A-Z]{2,4}\d{2}[A-Z]', line_text): continue  # E,NTR28E
                    if re.match(r'^[A-Z0-9,]+$', line_text) and len(line_text) < 30: continue  # all caps/nums (short)
                    # ── Hard skip: digit/date-only lines ────────────────
                    if re.match(r'^[\d/\s,\.]+$', line_text): continue
                    # ── Hard skip: sharer booking rows (room# + name starting with *) ─
                    if re.match(r'^\d{1,4}\s+\*', line_text): continue
                    # ── Hard skip: conf sub-rows (6+ digit number followed by space) ──
                    if re.match(r'^\d{6,}\s', line_text): continue
                    # ── Hard skip: raw rate/booking rows spilling over ───
                    if re.match(r'^\d[A-Z]\d[A-Z]{2}\s+\d', line_text): continue          # 1V5XK 0 0 0...
                    if re.search(r'\bDUO\b|\bOTA\b|\bDIU\b|\bEML\b', line_text) and re.search(r'\d{2}/\d{2}/\d{2}', line_text): continue
                    # ── Hard skip: fixed charges detail lines ────────────
                    if re.match(r'^\d{5}\s+(?:Guest|Adj)', line_text): continue            # 62500 Guest Transfer...
                    if re.match(r'^Guest\s+Transfer\s+\d', line_text): continue
                    # ── Hard skip: contract/board/commission noise ───────
                    if re.match(r'^(Contract\s+(Name|Description)|Board\s+description|'
                                 r'TA\s+(Commission|commission)|TA\s+commission|'
                                 r'Source\s+Channel|CONTRACTED\s+RATES|'
                                 r'Total\s+stays|Total\s+nights|PreRegistered|'
                                 r'Specials\s*:|Preferences\s*:|Membership\s+Type\s*:|'
                                 r'Promotions\s*:|Routing\s+Instruction|Fixed\s+Charges)',
                                 line_text, re.I): continue
                    # ── Hard skip: pure label lines with no content ──────
                    if re.match(r'^(Reservation\s+Notes|Reservation\s+Comment)\s*$',
                                 line_text, re.I): continue
                    # ── Hard skip: very short noise ──────────────────────
                    if len(line_text) <= 4: continue

                    # ── Extract content from label+content lines ─────────
                    # "Reservation Comment Dharsni - Allergic..." → keep content after label
                    m_label = re.match(
                        r'^(Reservation\s+(?:Notes|Comment)|Profile\s+Notes|'
                        r'Central\s+Comments|Reservation\s+Notes\s*:)\s*:?\s*(.+)$',
                        line_text, re.I)
                    if m_label:
                        content = m_label.group(2).strip()
                        # Skip if content is just another label
                        if re.match(r'^(Reservation\s+Comment|Total\s+stays|'
                                     r'Specials|Preferences)[\s:]', content, re.I):
                            continue
                        if len(content) >= 10:
                            note_lines.append(content)
                        continue

                    # ── Include: all remaining lines with real content ────
                    note_lines.append(line_text)
                # Check-in / Check-out dates and room type from room row
                checkin_w  = next((w for w in row if 288 <= w['x0'] <= 310
                                   and re.match(r'^\d{2}/\d{2}/\d{2}$', w['text'])), None)
                checkout_w = next((w for w in row if 326 <= w['x0'] <= 348
                                   and re.match(r'^\d{2}/\d{2}/\d{2}$', w['text'])), None)
                room_type_w = next((w for w in row if 364 <= w['x0'] <= 385
                                    and re.match(r'^1[A-Z]\d[A-Z]{2}$', w['text'])), None)
                adults_w  = next((w for w in row if 410 <= w['x0'] <= 432 and w['text'].isdigit()), None)
                child_w   = next((w for w in row if 434 <= w['x0'] <= 455 and w['text'].isdigit()), None)
                rate_w    = next((w for w in row if 560 <= w['x0'] <= 590
                                  and re.match(r'^[A-Z0-9]{6,}$', w['text'])), None)
                src_w     = next((w for w in row if 475 <= w['x0'] <= 515
                                  and re.match(r'^[A-Z]{2,4}$', w['text'])), None)

                checkin_str  = checkin_w['text']  if checkin_w  else ''
                checkout_str = checkout_w['text'] if checkout_w else ''
                room_type    = room_type_w['text'] if room_type_w else ''
                rate_code    = rate_w['text']     if rate_w     else ''
                src_code     = src_w['text']      if src_w      else ''
                adults_count  = int(adults_w['text'])  if adults_w  else 0
                child_count   = int(child_w['text'])   if child_w   else 0

                # Long stay flag — 7+ nights
                nights = 0
                if checkin_w and checkout_w:
                    try:
                        from datetime import datetime as _dt
                        ci = _dt.strptime(checkin_w['text'],  '%d/%m/%y')
                        co = _dt.strptime(checkout_w['text'], '%d/%m/%y')
                        nights = (co - ci).days
                        if nights >= 10:
                            bflags.append(f'Long Stay ({nights}N)')
                    except Exception:
                        pass

                # ── Potential LQA Auditor ─────────────────────────────────
                # Rate code starts with LQA prefix + direct booking + 3-4 nights
                _LQA_PREFIXES = ('S12','S14','M12','M14','O12','O14',
                                  'S16','M16','O16','O20','S23','S42','S43')
                _is_lqa_rate   = bool(rate_code and rate_code.startswith(_LQA_PREFIXES))
                _is_direct     = src_code in ('DIU','WTP','EML','WCO')
                _ta_prefix     = re.match(r'^([A-Z])-', ta_name)
                _not_wholesale = (_ta_prefix.group(1) != 'T') if _ta_prefix else True
                _is_lqa_stay   = nights in (3, 4)
                if _is_lqa_rate and _is_direct and _not_wholesale and _is_lqa_stay:
                    bflags.append('Potential LQA')

                # ── Payment Not Received ──────────────────────────────────
                # deposit=0.00 + not on credit list + not direct booking
                _is_direct_booking = bool(re.match(r'Direct|Complimentary', ta_name, re.I)
                                          or src_code in ('DIU', 'CMP'))
                # Comp = source code CMP, or standalone COMP (not "comp upgrade" from SPO)
                _is_comp = bool(src_code == 'CMP'
                                or re.search(r'(?<![A-Z/])\bCOMP\b(?!/)', ct, re.I)
                                or any(f.lower() == 'comp' for f in bflags))
                if deposit_val == 0.0 and not _on_credit and not _is_direct_booking and not _is_comp:
                    bflags.append('Payment Not Received')

                note = ' // '.join(note_lines[:15]) if note_lines else ''

                summary_guests.append({
                    'room':       rw['text'],
                    'name':       ' '.join(w['text'] for w in row if 41 <= w['x0'] <= 280)[:30],
                    'conf':       conf['text'] if conf else '',
                    'ta':         ta_name[:30],
                    'flight':     flight_str,
                    'arr_method': arr_method,
                    'flags':      bflags,
                    'note':       note,
                    'pdf_page':   pg_idx + 2,
                    'checkin':    checkin_str,
                    'checkout':   checkout_str,
                    'room_type':  room_type,
                    'rate_code':  rate_code,
                    'nights':     nights,
                    'adults':     adults_count,
                    'children':   child_count,
                    'deposit':    deposit_val,
                })

    # ══ PASS 4.5: orange annotation for bookings with no flight info ═════
    # flight_text is known per guest from Pass 4 — use it to annotate PDF
    if enabled_cats.get('flight'):
        no_flight_guests = [g for g in summary_guests if g['flight'] == 'NO FLIGHT INFO']
        for g in no_flight_guests:
            # Find the booking in all_bookings to get pg_idx and row_top
            bkg = next((b for b in all_bookings
                        if b['room'] == g['room'] and b['conf'] == g['conf']), None)
            if not bkg: continue
            pg_idx = bkg['pg_idx']
            t      = bkg['row_top']

            with pdfplumber.open(io.BytesIO(pdf_bytes)) as _pdf2:
                page  = _pdf2.pages[pg_idx]
                ph    = float(page.height)
                words = [w for w in page.extract_words(x_tolerance=2, y_tolerance=2)
                         if w['top'] < ph * 0.91]

            # Find conf line
            conf = next((w for w in words if w['top'] > t+3 and w['top'] < t+25
                         and 35 <= w['x0'] <= 180
                         and re.match(r'^[0-9]{6,}$', w['text'])), None)
            if not conf: continue

            conf_row = sorted([w for w in words if abs(w['top'] - conf['top']) <= 3],
                              key=lambda x: x['x0'])
            annot = make_annot(conf['x0'], conf['top'],
                               max(w['x1'] for w in conf_row), conf['bottom'],
                               ph, ORANGE, '⚠️ NO FLIGHT INFO — please check manually')
            p = writer.pages[pg_idx]
            if '/Annots' in p:
                p['/Annots'].append(annot)
            else:
                p[NameObject('/Annots')] = ArrayObject([annot])
            counts['flight'] += 1

    # Build summary data dict
    summary_data = {
        'property':  property_name,
        'date':      arrival_date,
        'generated': generated,
        'total':     len([b for b in all_bookings]),
        'rooms':     len(set(b['room'] for b in all_bookings if not b['is_sharer'])),
        'counts':    counts,
        'guests':    summary_guests,
    }

    # ══ PASS 5: build full report for HTML rendering ══════════════════════
    # Build lookup: (pg_idx, round(row_top)) → anchor id for each non-sharer booking
    booking_anchors = {}
    for g in summary_guests:
        for b in all_bookings:
            if b['room'] == g['room'] and b['conf'] == g['conf']:
                anchor = f"booking-{g['conf']}-{g['room']}"
                booking_anchors[(b['pg_idx'], round(b['row_top']))] = anchor
                break

    # Extract logo from first page
    import base64 as _b64
    logo_b64 = ''
    try:
        first_page = pdf.pages[0]
        fph = float(first_page.height)
        if first_page.images:
            img = first_page.images[0]
            top    = fph - img['y1']
            bottom = fph - img['y0']
            logo_crop = first_page.crop((img['x0'], top, img['x1'], bottom))
            logo_img_obj = logo_crop.to_image(resolution=200)
            import io as _io2
            logo_buf = _io2.BytesIO()
            logo_img_obj.save(logo_buf, format='PNG')
            logo_b64 = _b64.b64encode(logo_buf.getvalue()).decode()
    except Exception:
        pass

    full_report = []   # list of booking_blocks per page-group
    # We'll rebuild this as a list of booking blocks (not pages)
    # Each block = one booking's worth of lines, rendered as one unit

    # First collect ALL lines across all pages with their correct positions
    all_raw_lines = []  # {pg, top_abs, words, text, is_booking_start, anchor_id, ...}

    for pg_idx, page in enumerate(pdf.pages):
        ph = float(page.height)
        pw = float(page.width)
        footer_y = ph * 0.91

        # Use same tolerances as extract_text for faithful PDF reproduction
        # x_tolerance=3 groups chars into words correctly
        # y_tolerance=3 merges the split sub-rows (y=159+y=160) into one line
        words = [w for w in page.extract_words(
                     x_tolerance=3, y_tolerance=3,
                     extra_attrs=['fontname', 'size'])
                 if w['top'] < footer_y]

        # Same words for highlight detection
        hl_lines = defaultdict(list)
        for w in words:
            hl_lines[round(w['top'])].append(w)

        # Find header end
        header_end_top = 0
        for w in words:
            if w['text'] == 'Date' and w['x0'] < 80:
                row = [x for x in words if abs(x['top'] - w['top']) <= 4]
                if any(x['text'] == 'Arrival' for x in row):
                    header_end_top = round(w['top'])
                    break

        # Find booking start tops for this page
        booking_start_tops = set()
        for (bi, bt), anchor in booking_anchors.items():
            if bi == pg_idx:
                booking_start_tops.add(bt)

        # Detect highlighted lines
        highlighted_tops = {}
        for k, ws in hl_lines.items():
            ws_s = sorted(ws, key=lambda x: x['x0'])
            line_text = ' '.join(w['text'] for w in ws_s)
            if is_sharer_line(pg_idx, k): continue
            color = None
            for cat_id, cat in CATEGORIES.items():
                if not enabled_cats.get(cat_id, True): continue
                if cat.get('kw') and any(kw.lower() in line_text.lower() for kw in cat['kw']):
                    color = 'yellow'; break
            if color is None:
                if re.search(r'(EK\s*\d{3,4}|EY\s*\d{3,4}|QR\s*\d{3,4}|G9\s*\d{3,4}|KU\s*\d{3,4}|GF\s*\d{3,4})', line_text, re.I):
                    has_eta = bool(re.search(r'\b([0-1]?[0-9]|2[0-3]):[0-5][0-9]\b', line_text))
                    color = 'yellow' if has_eta else 'orange'
                elif re.search(r'\d+(st|nd|rd|th)\s+[Tt]ime\s+RPT|\bRPT\b', line_text): color = 'yellow'
                elif re.search(r'\d+(st|nd|rd|th)\s+[Ss]tay|Upcoming\s+[Ss]tay', line_text, re.I): color = 'yellow'
                elif re.search(r'Membership\s+Level\s+(GOLD|PLATINUM|TITANIUM|RED)', line_text, re.I): color = 'yellow'
                elif re.search(r'[1-9]\d*D\$', line_text, re.I): color = 'yellow'
                elif re.search(r'VIP[A-Z0-9]', line_text): color = 'yellow'
                elif re.search(r'Share\s+with:', line_text, re.I): color = 'yellow'
                elif re.search(r'0[0-8]:\d{2}', line_text) and not re.search(r'Arrival\s+Time', line_text, re.I): color = 'yellow'
                rw_m = next((w for w in ws_s if w['x0']<40 and re.match(r'^\d{1,4}$',w['text'])),None)
                if rw_m:
                    radl = next((w for w in ws_s if 415<=w['x0']<=435 and w['text'].isdigit()),None)
                    if radl and int(radl['text']) > 0: color = 'yellow'
            if color: highlighted_tops[k] = color

        # Group words into lines — y_tolerance=3 already merges sub-rows
        word_by_top = defaultdict(list)
        for w in words:
            word_by_top[round(w['top'])].append(w)

        for top_key in sorted(word_by_top.keys()):
            ws = sorted(word_by_top[top_key], key=lambda x: x['x0'])
            line_text = ' '.join(w['text'] for w in ws)
            top_r = top_key  # already rounded
            anchor_id = booking_anchors.get((pg_idx, top_r), '')
            if not anchor_id:
                for d in [1,-1,2,-2]:
                    anchor_id = booking_anchors.get((pg_idx, top_r+d), '')
                    if anchor_id: break

            is_hl = top_r in highlighted_tops
            if not is_hl:
                for d in [1,-1,2,-2]:
                    if top_r+d in highlighted_tops:
                        is_hl = True
                        highlighted_tops[top_r] = highlighted_tops[top_r+d]
                        break

            is_start = top_r in booking_start_tops
            if not is_start:
                for d in [1,-1]: 
                    if top_r+d in booking_start_tops: is_start=True; break

            word_positions = []
            for w in ws:
                fn = w.get('fontname','')
                sz = w.get('size', 8.0)
                bold = 'Bold' in fn and 'Italic' not in fn
                italic = 'Italic' in fn and 'Bold' not in fn
                bold_italic = 'Bold' in fn and 'Italic' in fn
                word_positions.append({
                    'text':   w['text'],
                    'x':      round(w['x0'], 1),
                    'sz':     round(sz, 1),
                    'bold':   bold or bold_italic,
                    'italic': italic or bold_italic,
                })

            all_raw_lines.append({
                'pg':               pg_idx,
                'top':              round(top_key, 1),
                'text':             line_text,
                'words':            word_positions,
                'highlighted':      is_hl,
                'color':            highlighted_tops.get(top_r, 'yellow'),
                'is_booking_start': is_start,
                'is_header_end':    abs(top_key - header_end_top) <= 1.5,
                'anchor_id':        anchor_id,
                'hl_id':            f"hl-{pg_idx}-{top_r}" if is_hl else '',
                'is_header_area':   top_key < 150,
            })

    # ── Now group all_raw_lines into booking BLOCKS ───────────────────────
    # Each block has a header section (page headers) + booking content lines
    # Header area = lines with top < 150 (page header repeated on every PDF page)
    # We collect ONE header block (from page 1) + individual booking blocks

    # Get the header lines (from first page only)
    header_lines = [l for l in all_raw_lines if l['pg'] == 0 and l['is_header_area']]

    # Identify booking block boundaries
    # A booking starts at is_booking_start=True
    booking_boundary_lines = [l for l in all_raw_lines if l['is_booking_start']]

    # Group all non-header lines into booking blocks
    # Each block = lines from one booking start to the next
    content_lines = [l for l in all_raw_lines if not l['is_header_area']]

    blocks = []  # list of {'header': [...], 'lines': [...]}

    # Add a pure header block first
    blocks.append({'is_header': True, 'lines': header_lines})

    # Split content lines into per-booking blocks
    current_block = []
    for line in content_lines:
        if line['is_booking_start'] and current_block:
            blocks.append({'is_header': False, 'lines': current_block})
            current_block = []
        current_block.append(line)
    if current_block:
        blocks.append({'is_header': False, 'lines': current_block})

    # Re-normalise top values within each block (relative to block start)
    for block in blocks:
        if block['is_header']:
            # Header: keep original tops
            for l in block['lines']:
                l['rel_top'] = l['top']
        else:
            lines = block['lines']
            if not lines: continue
            min_top = min(l['top'] for l in lines)
            for l in lines:
                l['rel_top'] = round(l['top'] - min_top, 1)

    # Merge consecutive booking blocks with same anchor (caused by TA line at y-1)
    merged_blocks = []
    for block in blocks:
        if (not block.get('is_header') and merged_blocks and 
            not merged_blocks[-1].get('is_header')):
            # Get anchor of last block and this block
            def get_anchor(b):
                for l in b['lines']:
                    if l.get('anchor_id'): return l['anchor_id']
                return None
            last_anchor = get_anchor(merged_blocks[-1])
            this_anchor = get_anchor(block)
            if last_anchor and last_anchor == this_anchor:
                # Merge: extend lines and recalculate rel_top
                merged_blocks[-1]['lines'].extend(block['lines'])
                # Renormalize rel_top
                all_block_lines = merged_blocks[-1]['lines']
                min_top = min(l['top'] for l in all_block_lines)
                for l in all_block_lines:
                    l['rel_top'] = round(l['top'] - min_top, 1)
                continue
        merged_blocks.append(block)
    blocks = merged_blocks

    summary_data['full_report_blocks'] = blocks
    summary_data['logo_b64']           = logo_b64


    # ── Pre-assign synthetic hl_ids to anchor lines ──────────────────────
    # Must happen BEFORE page images so overlays are emitted correctly
    for block in blocks:
        for line in block['lines']:
            if line.get('anchor_id') and not line.get('hl_id'):
                line['hl_id'] = f"anchor-{line['anchor_id']}"

    # ── PASS 6: render each PDF page as image with highlights ────────────
    try:
        from PIL import Image as _PILImage, ImageDraw as _PILDraw
        import base64 as _b64mod

        DPI   = 120
        SCALE = DPI / 72.0  # 1pt = 1.667px at 120dpi

        page_images = []  # list of {'pg', 'b64', 'width_px', 'height_px', 'anchors'}
        for pg_idx, page in enumerate(pdf.pages):
            ph = float(page.height)
            pw = float(page.width)

            # Render clean page — no highlight overlay
            # Interactive glow handles highlighting, PDF has the yellow
            img_obj  = page.to_image(resolution=DPI)
            pil_img  = img_obj.original.convert('RGB')
            iw, ih   = pil_img.size
            buf = io.BytesIO()
            pil_img.save(buf, format='JPEG', quality=75)
            b64 = _b64mod.b64encode(buf.getvalue()).decode()

            # Collect anchor positions — one per booking, use the room/name line (x0<40)
            anchors = []
            seen_anchor_ids = set()
            for block in blocks:
                for line in block['lines']:
                    aid = line.get('anchor_id', '')
                    if not aid or line.get('pg') != pg_idx:
                        continue
                    if aid in seen_anchor_ids:
                        continue
                    # Prefer the room/name line (has word at x0<40) over the TA line
                    words = line.get('words', [])
                    is_room_line = any(w['x'] < 40 for w in words)
                    if not is_room_line and any(
                        l.get('anchor_id') == aid and l.get('pg') == pg_idx
                        and any(w['x'] < 40 for w in l.get('words', []))
                        for l in (ll for b in blocks for ll in b['lines'])
                    ):
                        continue  # skip TA line, room line will be processed
                    seen_anchor_ids.add(aid)
                    anchors.append({
                        'id':    aid,
                        'y_px':  int(line['top'] * SCALE),
                        'hl_id': line.get('hl_id', ''),
                    })

            page_images.append({
                'pg':       pg_idx,
                'b64':      b64,
                'width_px': iw,
                'height_px': ih,
                'anchors':  anchors,
                'scale':    SCALE,
            })

        summary_data['page_images'] = page_images
    except Exception as e:
        summary_data['page_images'] = []

    # Keep full_report for backward compat (empty)
    full_report = []
    summary_data['full_report'] = full_report

    # Build summary PDF page and prepend to highlighted PDF
    summary_bytes = build_summary_page(summary_data)

    # Merge: summary page first, then highlighted report
    final_writer = PdfWriter()
    summary_reader = PdfReader(io.BytesIO(summary_bytes))
    for p in summary_reader.pages:
        final_writer.add_page(p)
    highlighted_reader = PdfReader(io.BytesIO((lambda b: (writer.write(b), b)[1])(io.BytesIO())[1] if False else _get_writer_bytes(writer)))
    for p in highlighted_reader.pages:
        final_writer.add_page(p)

    out = io.BytesIO()
    final_writer.write(out)
    return out.getvalue(), counts, total, summary_data


def _get_writer_bytes(writer):
    buf = io.BytesIO()
    writer.write(buf)
    buf.seek(0)
    return buf.read()


def build_summary_html(summary_data, pdf_filename='highlighted.pdf'):
    """Build interactive HTML: summary table + full rendered report with highlights."""
    import json, html as htmllib

    guests = summary_data['guests']
    counts = summary_data['counts']

    # ── Map flags to filter categories ───────────────────────────────────
    guests_json = []
    blocks_data = summary_data.get('full_report_blocks', [])
    # Flatten all lines from booking blocks for hl lookup
    all_lines_flat = []
    for block in blocks_data:
        all_lines_flat.extend(block['lines'])

    # Build booking→highlighted lines map
    def get_hl_lines_for_guest(g):
        anchor = f"booking-{g['conf']}-{g['room']}"
        result = {}

        # Find the booking block that has this anchor
        booking_block = None
        for block in blocks_data:
            if not block.get('is_header'):
                for line in block['lines']:
                    if line.get('anchor_id') == anchor:
                        booking_block = block
                        break
            if booking_block: break

        if not booking_block:
            return result

        # Always include the booking header lines under 'booking'
        # Use the room/name line (x0 < 40) as the primary anchor — hl_id already pre-assigned
        anchor_lines = [l for l in booking_block['lines'] if l.get('anchor_id') == anchor]
        # Prefer room/name line (has word at x<40), fall back to any anchor line
        room_lines = [l for l in anchor_lines if any(w['x'] < 40 for w in l.get('words', []))]
        primary_lines = room_lines if room_lines else anchor_lines[:1]
        booking_hl_ids = []
        seen_ids = set()
        for l in primary_lines:
            hid = l.get('hl_id') or f"anchor-{anchor}"
            l['hl_id'] = hid
            if hid not in seen_ids:
                booking_hl_ids.append(hid)
                seen_ids.add(hid)
        if booking_hl_ids:
            result['booking'] = booking_hl_ids

        CAT_KW = {
            'membership': ['membership level','platinum','titanium','gold','red','silver'],
            'vip':        ['vip'],
            'repeater':   ['rpt','repeat','time rpt','stays','1st time','2nd time','3rd time','4th time','5th time'],
            'allergy':    ['allerg','shellfish','gluten','peanut','lactose','vegan','vegetarian','halal','intoleran'],
            'complaint':  ['complaint','glitch','recovery','upset','feedback','inconveni'],
            'payment':    ['collect upon arrival','to be collected','pls collect','please collect','balance due','poa'],
            'occasion':   ['honeymoon','anniversary','birthday','wedding','babymoon','proposal','celebrating','engagement'],
            'together':   ['share with','together','t#','party'],
            'legs':       ['leg','1st leg','2nd leg'],
            'dbalance':   ['d$'],
            'flight':     ['ek ','ey ','qr ','g9 ','ku ','gf ','sq ','ai ','6e ','ek\t','ey\t'],
            'eci':        ['eci'],
            'lco':        ['lco'],
            'upgrade':    ['upg','upgrade'],
            'comp':       ['comp'],
            'children':   ['child','chl','infant','baby','kids'],
            'noflight':   ['no flight info'],
            'longstay':   ['long stay'],
            'potvip':     ['ceo','coo','cfo','hrd','hrh','vice president','managing director',
                           'ambassador','prime minister','governor','senator','his excellency',
                           'her excellency','royal household','embassy of','ministry of'],
            'restaurant': ['fushi','sea fire salt','sfs','aqua','yellow fin','origami',
                           'dhoni bar','cumin','baan huraa','designer dining','living room'],
            'spa':        ['anantara spa','dhigu spa','veli spa','spa booking','spa reserv',
                           'spa treatment','couples massage','four hands massage'],
            'activity':   ['snorkel','scuba','diving','sunset cruise','fishing trip',
                           'dolphin','whale','kayak','watersport','sandbank','overwater dinner'],
            'sharer':     ['share with:'],
            'poa':        ['poa/','poa ','collect upon arrival'],
        }

        # Assign synthetic hl_ids to non-highlighted lines for all flag categories
        new_flag_kws = {
            'eci':        ['eci'],
            'lco':        ['lco'],
            'upgrade':    ['upg','upgrade'],
            'comp':       ['comp'],
            'children':   ['child','chl'],
            'dbalance':   ['d$'],
            'noflight':   ['no flight'],
            'longstay':   ['long stay'],
            'restaurant': ['fushi','sea fire salt','origami','dhoni bar','cumin',
                           'baan huraa','designer dining','living room','yellow fin','aqua','sfs'],
            'spa':        ['anantara spa','dhigu spa','veli spa','spa booking','spa reserv',
                           'spa treatment','couples massage','four hands massage'],
            'activity':   ['snorkel','scuba','diving','sunset cruise','fishing trip',
                           'dolphin','whale','kayak','watersport','sandbank'],
            'potvip':     ['vice president','managing director','ceo','coo','cfo',
                           'ambassador','prime minister','his excellency','her excellency',
                           'royal household','embassy of','ministry of'],
            'poa':        ['poa/','poa '],
            'sharer':     ['share with:'],
        }
        for line in booking_block['lines']:
            if line.get('hl_id'):
                continue  # already has one
            txt = line['text'].lower()
            for cat, kws in new_flag_kws.items():
                if any(kw in txt for kw in kws):
                    # Give it a synthetic hl_id so JS can find and flash it
                    synth_id = f"flag-{cat}-{line['pg']}-{int(line['top'])}"
                    line['hl_id'] = synth_id
                    break  # one id per line is enough

        for line in booking_block['lines']:
            if not line.get('hl_id'):
                continue
            txt = line['text'].lower()
            for cat, kws in CAT_KW.items():
                if any(kw in txt for kw in kws):
                    result.setdefault(cat, list(booking_hl_ids))  # always start with booking lines
                    if line['hl_id'] not in result[cat]:
                        result[cat].append(line['hl_id'])
            # Flight lines (airline codes without space after)
            if re.search(r'\b(EK|EY|QR|G9|KU|GF|SQ|AI|6E)\s*\d{3,4}\b', line['text'], re.I):
                result.setdefault('flight', list(booking_hl_ids))
                if line['hl_id'] not in result['flight']:
                    result['flight'].append(line['hl_id'])

        # For every category, prepend the booking header lines
        for cat in list(result.keys()):
            if cat != 'booking':
                for bid in reversed(booking_hl_ids):
                    if bid not in result[cat]:
                        result[cat].insert(0, bid)

        return result

    for g in guests:
        cats = []
        for f in g['flags']:
            fl = f.lower()
            if any(x in fl for x in ['platinum','titanium','gold','red','silver']): cats.append('membership')
            if 'vip' in fl: cats.append('vip')
            if any(x in fl for x in ['rpt','repeat','3rd time','2nd time','4th time','5th time']): cats.append('repeater')
            if any(x in fl for x in ['allerg','shellfish','gluten','peanut','lactose','vegan','halal','intoleran']): cats.append('allergy')
            if 'complaint' in fl or 'glitch' in fl: cats.append('complaint')
            if 'collect' in fl or 'payment' in fl: cats.append('payment')
            if any(x in fl for x in ['honeymoon','anniversary','birthday','wedding','babymoon','proposal','engagement']): cats.append('occasion')
            if 'together' in fl: cats.append('together')
            if 'multi-leg' in fl or ('leg' in fl and 'multi' in fl): cats.append('legs')
            if 'd$' in fl: cats.append('dbalance')
            if 'early check-in' in fl: cats.append('eci')
            if 'late check-out' in fl: cats.append('lco')
            if 'upgrade' in fl: cats.append('upgrade')
            if fl == 'comp': cats.append('comp')
            if 'child' in fl: cats.append('children')
            if 'long stay' in fl: cats.append('longstay')
            if 'potential vip' in fl: cats.append('potvip')
            if 'restaurant' in fl: cats.append('restaurant')
            if 'spa' in fl: cats.append('spa')
            if 'activit' in fl: cats.append('activity')
            if 'sharer missing' in fl: cats.append('sharermissing')
            if 'poa:' in fl or fl == 'poa amount': cats.append('poa')
            if 'potential lqa' in fl: cats.append('lqa')
            if 'payment not received' in fl: cats.append('paymissing')
        if g['flight'] and 'NO ETA' in g['flight']: cats.append('noeta')
        if g['flight'] and g['flight'] not in ('', '--', '-', 'NO FLIGHT INFO'): cats.append('flight')
        if g['flight'] == 'NO FLIGHT INFO': cats.append('noflight')

        hl_lines = get_hl_lines_for_guest(g)

        guests_json.append({
            'room':       g['room'], 'name': g['name'], 'conf': g['conf'],
            'ta':         g.get('ta',''),
            'flight':     g['flight'],
            'arr_method': g.get('arr_method',''),
            'flags':      g['flags'],
            'note':       g['note'], 'cats': list(set(cats)),
            'anchor':     f"booking-{g['conf']}-{g['room']}",
            'hl_lines':   hl_lines,
            'checkin':    g.get('checkin',''),
            'checkout':   g.get('checkout',''),
            'room_type':  g.get('room_type',''),
            'rate_code':  g.get('rate_code',''),
            'nights':     g.get('nights', 0),
            'adults':     g.get('adults', 0),
            'children':   g.get('children', 0),
            'deposit':    g.get('deposit', 0.0),
        })

    # ── Stat cards ────────────────────────────────────────────────────────
    # Count unique rooms per category (not PAX / highlight count)
    def count_rooms(cat_key):
        seen = set()
        for g in guests:
            cats_g = []
            for f in g['flags']:
                fl = f.lower()
                if any(x in fl for x in ['platinum','titanium','gold','red','silver']): cats_g.append('membership')
                if 'vip' in fl: cats_g.append('vip')
                if any(x in fl for x in ['rpt','repeat','time rpt']): cats_g.append('repeater')
                if any(x in fl for x in ['allerg','shellfish','gluten','peanut','lactose','vegan','halal']): cats_g.append('allergy')
                if 'complaint' in fl or 'glitch' in fl: cats_g.append('complaint')
                if 'collect' in fl or 'payment' in fl: cats_g.append('payment')
                if any(x in fl for x in ['honeymoon','anniversary','birthday','wedding','babymoon','proposal','engagement']): cats_g.append('occasion')
                if 'together' in fl: cats_g.append('together')
                if 'multi-leg' in fl or ('leg' in fl and 'multi' in fl): cats_g.append('legs')
                if 'd$' in fl: cats_g.append('dbalance')
                if 'early check-in' in fl: cats_g.append('eci')
                if 'late check-out' in fl: cats_g.append('lco')
                if 'upgrade' in fl: cats_g.append('upgrade')
                if fl == 'comp': cats_g.append('comp')
                if 'child' in fl: cats_g.append('children')
                if 'long stay' in fl: cats_g.append('longstay')
                if 'potential vip' in fl: cats_g.append('potvip')
                if 'restaurant' in fl: cats_g.append('restaurant')
                if 'spa' in fl: cats_g.append('spa')
                if 'activit' in fl: cats_g.append('activity')
                if 'sharer missing' in fl: cats_g.append('sharermissing')
                if 'poa:' in fl or fl == 'poa amount': cats_g.append('poa')
                if 'potential lqa' in fl: cats_g.append('lqa')
                if 'payment not received' in fl: cats_g.append('paymissing')
            if g.get('flight') and g['flight'] not in ('','--','-','NO FLIGHT INFO'): cats_g.append('flight')
            if g.get('flight') and 'NO ETA' in g.get('flight',''): cats_g.append('noeta')
            if g.get('flight') == 'NO FLIGHT INFO': cats_g.append('noflight')
            if cat_key in cats_g:
                seen.add(g['room'])
        return len(seen)

    stat_items = [
        ('rooms',      '🏠', 'Rooms',      summary_data['rooms'],      'all'),
        ('repeater',   '⭐', 'Repeaters',  count_rooms('repeater'),    'repeater'),
        ('flight',     '✈️', 'Flights',    count_rooms('flight'),      'flight'),
        ('noflight',   '❓', 'No Flight',  len([g for g in guests if g.get('flight') == 'NO FLIGHT INFO']), 'noflight'),
        ('membership', '👑', 'GHA',        count_rooms('membership'),  'membership'),
        ('vip',        '💎', 'VIP',        count_rooms('vip'),         'vip'),
        ('complaint',  '⚠️', 'Complaints', count_rooms('complaint'),   'complaint'),
        ('allergy',    '🌿', 'Allergies',  count_rooms('allergy'),     'allergy'),
        ('occasion',   '🎂', 'Occasions',  count_rooms('occasion'),    'occasion'),
        ('payment',    '💳', 'Payment',    count_rooms('payment'),     'payment'),
        ('dbalance',   '💰', 'D$ Balance', count_rooms('dbalance'),    'dbalance'),
        ('together',   '👥', 'Together',   len(set(g['room'] for g in guests if any('Together' in f for f in g['flags']))), 'together'),
        ('eci',        '🌅', 'Early CI',   len([g for g in guests if any('Early Check-in' in f for f in g['flags'])]), 'eci'),
        ('lco',        '🌙', 'Late CO',    len([g for g in guests if any('Late Check-out' in f for f in g['flags'])]), 'lco'),
        ('upgrade',    '⬆️', 'Upgrade',    len([g for g in guests if any('Upgrade' in f for f in g['flags'])]), 'upgrade'),
        ('comp',       '🎁', 'Comp',       len([g for g in guests if any('Comp' in f for f in g['flags'])]), 'comp'),
        ('children',   '👶', 'Children',   len([g for g in guests if any('Child' in f for f in g['flags'])]), 'children'),
        ('longstay',   '🌴', 'Long Stay',  len([g for g in guests if any('Long Stay' in f for f in g['flags'])]), 'longstay'),
        ('potvip',     '⚑',  'Pot. VIP',   len([g for g in guests if any('Potential VIP' in f for f in g['flags'])]), 'potvip'),
        ('restaurant', '🍽️', 'Restaurant', len([g for g in guests if any('Restaurant' in f for f in g['flags'])]), 'restaurant'),
        ('spa',        '💆', 'Spa',        len([g for g in guests if any('Spa' in f for f in g['flags'])]), 'spa'),
        ('activity',   '🤿', 'Activities', len([g for g in guests if any('Activit' in f for f in g['flags'])]), 'activity'),
        ('sharermissing','⚠️','Sharer?',   len([g for g in guests if any('Sharer Missing' in f for f in g['flags'])]), 'sharermissing'),
        ('poa',        '💵', 'POA Amt',    len([g for g in guests if any('POA' in f for f in g['flags'])]), 'poa'),
        ('lqa',        '🔍', 'Pot. LQA',   len([g for g in guests if any('Potential LQA' in f for f in g['flags'])]), 'lqa'),
        ('paymissing', '🚨', 'Pay Missing', len([g for g in guests if any('Payment Not Received' in f for f in g['flags'])]), 'paymissing'),
    ]
    stat_cards = ''
    for sid, icon, label, val, cat in stat_items:
        active_cls = 'active' if cat == 'all' else ''
        stat_cards += f'<div class="stat-card {active_cls}" onclick="filterGuests(\'{cat}\')" data-cat="{cat}"><div class="stat-icon">{icon}</div><div class="stat-num">{val}</div><div class="stat-label">{label}</div></div>'

    # ── Full report HTML rendering — images of actual PDF pages ─────────
    report_html = ''
    page_images  = summary_data.get('page_images', [])
    emitted_anchor_ids = set()   # prevent duplicate id= attributes
    emitted_hlids      = set()   # prevent duplicate data-hlid= attributes

    for pi in page_images:
        b64  = pi['b64']
        iw   = pi['width_px']
        ih   = pi['height_px']
        pg   = pi['pg']
        sc   = pi.get('scale', 1.667)

        report_html += (f'<div class="report-page" id="rpage-{pg+1}" '
                        f'style="width:{iw}px;height:{ih}px;position:relative">')
        report_html += (f'<img src="data:image/jpeg;base64,{b64}" '
                        f'width="{iw}" height="{ih}" style="display:block;position:absolute;top:0;left:0">')

        # Invisible anchor divs — emit each anchor ID only once
        for anc in pi.get('anchors', []):
            aid   = anc['id']
            y_px  = anc['y_px']
            hl_id = anc.get('hl_id', '')
            if aid in emitted_anchor_ids:
                continue
            emitted_anchor_ids.add(aid)
            hl_attr = f' data-hlid="{hl_id}"' if hl_id and hl_id not in emitted_hlids else ''
            if hl_id:
                emitted_hlids.add(hl_id)
            report_html += (f'<div id="{aid}" class="booking-anchor"{hl_attr} '
                            f'style="position:absolute;top:{y_px}px;left:0;'
                            f'width:100%;height:30px;pointer-events:none"></div>')

        # hl_id overlays for flash — emit each hlid only once
        for block in blocks_data:
            for line in block['lines']:
                if not line.get('hl_id') or line.get('pg') != pg:
                    continue
                hlid = line['hl_id']
                if hlid in emitted_hlids:
                    continue
                emitted_hlids.add(hlid)
                top_px = int(line['top'] * sc)
                line_h = int(10 * sc)  # ~16px at 120dpi
                report_html += (f'<div data-hlid="{hlid}" '
                                f'style="position:absolute;top:{top_px}px;left:0;'
                                f'width:100%;height:{line_h}px;pointer-events:none"></div>')

        report_html += '</div>'

    guests_js = json.dumps(guests_json)

    # Build GHA level checkboxes
    _GHA_LEVELS = [
        ('Silver',   '#94a3b8'),
        ('Gold',     '#b7960c'),
        ('Red',      '#dc2626'),
        ('Platinum', '#6366f1'),
        ('Titanium', '#0f766e'),
    ]
    gha_checkboxes = ''.join(
        f'<label style="display:flex;align-items:center;gap:3px;font-size:10px;'
        f'cursor:pointer;user-select:none;white-space:nowrap">'
        f'<input type="checkbox" class="gha-level-toggle" data-level="{lvl.lower()}" checked '
        f'onchange="updateGHACount();render(currentFilter)" '
        f'style="accent-color:{col};width:12px;height:12px;cursor:pointer">'
        f'<span style="color:{col};font-weight:600">{lvl}</span></label>'
        for lvl, col in _GHA_LEVELS
    )

    # ── Payment missing summary data ─────────────────────────────────────
    pay_missing = [g for g in guests if any('Payment Not Received' in f for f in g['flags'])]
    pay_missing_js = json.dumps([{
        'room':      g['room'],
        'name':      g['name'],
        'conf':      g['conf'],
        'ta':        g.get('ta',''),
        'checkin':   g.get('checkin',''),
        'checkout':  g.get('checkout',''),
        'nights':    g.get('nights',0),
        'rate_code': g.get('rate_code',''),
        'deposit':   g.get('deposit',0.0),
        'anchor':    f"booking-{g['conf']}-{g['room']}",
    } for g in pay_missing])

    html = f'''<!DOCTYPE html>
<html lang="en">
<head>
<meta charset="UTF-8">
<meta name="viewport" content="width=device-width,initial-scale=1">
<title>Daily Operations Meeting — {summary_data["date"]}</title>
<style>
*{{box-sizing:border-box;margin:0;padding:0}}
body{{font-family:-apple-system,BlinkMacSystemFont,'Segoe UI',sans-serif;background:#f1f5f9;color:#1e2535;font-size:13px}}

/* ── Header ── */
.header{{background:#1a1a2e;color:#fff;padding:14px 24px;display:flex;align-items:center;justify-content:space-between;position:sticky;top:0;z-index:100}}
.header h1{{font-size:17px;font-weight:700}}
.header-sub{{font-size:11px;color:#94a3b8;margin-top:2px}}
.property{{font-size:17px;font-weight:700;color:#fff;text-align:right}}
.nav-tabs{{display:flex;gap:0;background:#f1f5f9;border-bottom:2px solid #e2e8f0}}
.nav-tab{{padding:10px 28px;font-size:13px;font-weight:600;color:#6b7280;cursor:pointer;border-bottom:2px solid transparent;margin-bottom:-2px;transition:all .15s;user-select:none}}
.nav-tab.active{{color:#1a56db;border-bottom-color:#1a56db;background:#fff}}
.nav-tab:hover:not(.active){{color:#374151;background:#e9ecef}}

/* ── Stats ── */
.stats{{display:flex;gap:8px;padding:12px 24px;background:#fff;border-bottom:1px solid #e2e8f0;flex-wrap:wrap}}
.stat-card{{flex:1;min-width:72px;background:#f8fafc;border:1.5px solid #e2e8f0;border-radius:10px;padding:8px 6px;text-align:center;cursor:pointer;transition:all .15s;user-select:none}}
.stat-card:hover{{border-color:#94a3b8;background:#f1f5f9}}
.stat-card.active{{background:#1a56db;border-color:#1a56db}}
.stat-card.active .stat-num,.stat-card.active .stat-label,.stat-card.active .stat-icon{{color:#fff !important}}
.stat-icon{{font-size:15px;margin-bottom:2px}}
.stat-num{{font-size:20px;font-weight:700;color:#1a1a2e;line-height:1}}
.stat-label{{font-size:10px;color:#6b7280;margin-top:2px;font-weight:500}}
.filter-bar{{padding:8px 24px;background:#fff;border-bottom:1px solid #e2e8f0;display:flex;align-items:center;gap:8px}}
.filter-label{{font-size:11px;color:#6b7280;font-weight:600;text-transform:uppercase;letter-spacing:.05em}}
.filter-tag{{font-size:11px;background:#1a56db;color:#fff;padding:3px 10px;border-radius:20px}}
.count-badge{{font-size:11px;color:#6b7280}}

/* ── Summary table ── */
.table-wrap{{padding:16px 24px}}
table{{width:100%;border-collapse:collapse;background:#fff;border-radius:10px;overflow:hidden;box-shadow:0 1px 3px rgba(0,0,0,.06)}}
th{{background:#f8fafc;font-size:10px;font-weight:700;text-transform:uppercase;letter-spacing:.07em;color:#6b7280;padding:9px 11px;text-align:left;border-bottom:1px solid #e2e8f0}}
th:last-child{{width:45%}}
td{{padding:9px 11px;border-bottom:1.5px solid #e2e8f0;vertical-align:top}}
tbody tr:nth-child(even){{background:#fafbfd}}
tbody tr:hover{{background:#f0f7ff!important}}
td:last-child{{width:40%}}
tr:last-child td{{border-bottom:none}}
tr.dimmed{{opacity:.2;transition:opacity .2s}}
.room-no{{font-weight:700;font-size:14px}}
.guest-name{{font-weight:500}}
.guest-link{{cursor:pointer;color:#1a56db;text-decoration:underline;text-underline-offset:2px}}
.guest-link:hover{{color:#1040b0}}
.ta-name{{font-size:11px;color:#6b7280;font-style:italic}}
.conf-no{{font-size:11px;color:#94a3b8;font-family:monospace}}
.flight-badge{{display:inline-block;font-size:11px;padding:2px 8px;border-radius:12px;background:#FAEEDA;color:#633806;font-weight:500}}
.flight-badge.noeta{{background:#E24B4A;color:#fff}}
.flags{{display:flex;flex-wrap:wrap;gap:4px}}
.flag{{display:inline-block;font-size:10.5px;padding:2px 9px;border-radius:12px;font-weight:500;cursor:pointer;transition:opacity .15s}}
.flag:hover{{opacity:.75}}
/* ── Notes ── */
.note-original-text{{font-size:11.5px;color:#374151;line-height:1.55;white-space:pre-wrap;word-break:break-word;padding:3px 0 4px 0;min-height:16px;font-family:-apple-system,BlinkMacSystemFont,'Segoe UI',sans-serif;cursor:text;user-select:text}}
.note-hl{{background:#FFF176;border-radius:2px;padding:0 1px}}
.note-hl-toolbar{{position:absolute;z-index:500;background:#1a1a2e;border-radius:8px;padding:4px 8px;display:flex;gap:6px;align-items:center;box-shadow:0 4px 14px rgba(0,0,0,.35);white-space:nowrap}}
.note-hl-toolbar button{{background:none;border:none;cursor:pointer;font-size:13px;padding:2px 5px;border-radius:4px;color:#fff;transition:background .12s}}
.note-hl-toolbar button:hover{{background:rgba(255,255,255,.15)}}
.note-manual-field{{font-size:12px;font-family:Georgia,serif;font-style:italic;color:#b91c1c;width:100%;min-height:26px;padding:3px 6px;border:1px dashed #fca5a5;border-radius:4px;background:#fff5f5;resize:vertical;outline:none;line-height:1.5;transition:border .15s;margin-top:3px;display:block}}
.note-manual-field:focus{{border-color:#ef4444;background:#fff}}
.note-manual-field::placeholder{{color:#fca5a5;font-style:italic;font-family:Georgia,serif}}
.dl-pdf-btn{{margin-left:auto;font-size:11px;font-weight:600;color:#1a1a2e;background:#f1f5f9;border:1px solid #cbd5e1;border-radius:20px;padding:4px 14px;cursor:pointer;transition:background .15s}}
.dl-pdf-btn:hover{{background:#e2e8f0}}
.no-results{{text-align:center;padding:48px;color:#94a3b8;font-size:14px}}

/* ── Full report — PDF-style layout ── */
#section-report{{display:none;padding:20px 24px;background:#525659;overflow-x:auto;zoom:110%}}
#section-summary{{zoom:100%}}
.report-page{{position:relative;margin:0 auto 16px auto;box-shadow:0 4px 16px rgba(0,0,0,.4);display:block;overflow:hidden}}
.booking-anchor{{scroll-margin-top:80px}}
@keyframes bookingPulse{{
  0%   {{box-shadow: 4px 0 0 0 #d4a017 inset; background:rgba(255,245,80,0.18)}}
  50%  {{box-shadow: 4px 0 0 0 #d4a017 inset, 0 0 20px 6px rgba(255,235,50,0.50); background:rgba(255,245,80,0.55)}}
  100% {{box-shadow: 4px 0 0 0 #d4a017 inset; background:rgba(255,245,80,0.18)}}
}}
.booking-flash{{animation:bookingPulse 1.8s ease-in-out infinite;z-index:10}}
.hl-line-flash{{animation:bookingPulse 1.8s ease-in-out infinite;z-index:11;box-shadow:3px 0 0 0 #c0610a inset}}
/* Floating back-to-summary button */
.float-back{{position:fixed;bottom:28px;right:28px;z-index:999;background:#1a1a2e;color:#fff;border:none;border-radius:30px;padding:11px 20px;font-size:13px;font-weight:600;cursor:pointer;box-shadow:0 4px 16px rgba(0,0,0,.35);display:none;align-items:center;gap:7px;transition:background .15s}}
.float-back:hover{{background:#2d2d50}}
.float-back.visible{{display:flex}}

/* ── Mobile responsive ── */
@media (max-width:768px){{
  .header{{padding:10px 14px;flex-wrap:wrap;gap:4px}}
  .header h1{{font-size:14px}}
  .property{{font-size:13px;text-align:left;width:100%}}
  .header-sub{{font-size:10px}}
  .nav-tab{{padding:8px 16px;font-size:12px}}
  .stats{{padding:8px 12px;gap:6px}}
  .stat-card{{min-width:60px;padding:6px 4px}}
  .stat-num{{font-size:16px}}
  .stat-label{{font-size:9px}}
  .stat-icon{{font-size:13px}}
  .filter-bar{{padding:6px 12px;flex-wrap:wrap;gap:6px}}
  .dl-pdf-btn{{margin-left:0}}
  .table-wrap{{padding:8px 0}}
  table{{font-size:12px;border-radius:0}}
  th{{padding:7px 8px;font-size:9px}}
  td{{padding:7px 8px}}
  td:last-child{{min-width:0;width:auto}}
  th:last-child{{width:auto}}
  .room-no{{font-size:13px}}
  .note-manual-field{{font-size:11px}}
  /* Stack table horizontally scrollable on small screens */
  .table-wrap{{overflow-x:auto;-webkit-overflow-scrolling:touch}}
  table{{min-width:600px}}
  #section-report{{padding:8px;zoom:100%}}
  .float-back{{bottom:16px;right:16px;padding:9px 16px;font-size:12px}}
}}
</style>
</head>
<body>

<div class="header">
  <div>
    <h1>Daily Operations Meeting</h1>
    <div class="header-sub">{summary_data["date"]} &nbsp;·&nbsp; {summary_data["total"]} arrivals &nbsp;·&nbsp; Generated: {summary_data["generated"]}</div>
  </div>
  <div class="property">{summary_data["property"]}</div>
</div>

<div class="nav-tabs">
  <div class="nav-tab active" id="tab-summary" onclick="showSection('summary')">📋 Summary</div>
  <div class="nav-tab" id="tab-report" onclick="showSection('report')">📄 Full Report</div>
  <div class="nav-tab" id="tab-payment" onclick="showSection('payment')">🚨 Payment Missing <span id="pay-count-badge" style="background:#E24B4A;color:#fff;border-radius:10px;padding:1px 7px;font-size:10px;margin-left:4px">{len(pay_missing)}</span></div>
</div>

<div id="section-summary">
  <div class="stats">{stat_cards}</div>
  <div class="filter-bar">
    <span class="filter-label">Showing:</span>
    <span class="filter-tag" id="filter-tag">All Bookings</span>
    <span class="count-badge" id="count-badge">{len(guests)} guests</span>
    <div style="display:flex;align-items:center;gap:4px;margin-left:8px;flex-wrap:wrap">
      <span style="font-size:10px;color:#6b7280;font-weight:600;text-transform:uppercase;letter-spacing:.04em;margin-right:2px">GHA:</span>
      {gha_checkboxes}
    </div>
    <button class="dl-pdf-btn" onclick="downloadPDF()" title="Download summary as PDF">⬇️ Download PDF</button>
  </div>
  <div class="table-wrap">
    <table id="main-table">
      <thead>
        <tr><th>Room</th><th>Guest</th><th>Flight</th><th>Flags</th><th>Notes <span style="font-size:9px;font-weight:400;color:#94a3b8">(auto-saved · editable)</span></th></tr>
      </thead>
      <tbody id="tbody"></tbody>
    </table>
    <div class="no-results" id="no-results" style="display:none">No bookings match this filter</div>
  </div>
</div>

<div id="section-report">
{report_html}
</div>

<div id="section-payment" style="display:none;padding:20px 24px">
  <div style="background:#fff;border-radius:10px;box-shadow:0 1px 3px rgba(0,0,0,.06);overflow:hidden">
    <div style="background:#1a1a2e;color:#fff;padding:14px 20px;display:flex;align-items:center;gap:12px;flex-wrap:wrap">
      <div style="display:flex;align-items:center;gap:10px;flex:1;min-width:0">
        <span style="font-size:20px">🚨</span>
        <div>
          <div style="font-weight:700;font-size:15px">Payment Not Received</div>
          <div style="font-size:11px;color:#94a3b8;margin-top:2px">Zero deposit · Not on credit account · Follow up required</div>
        </div>
      </div>
      <div style="display:flex;align-items:center;gap:8px;flex-shrink:0">
        <span id="pay-total-badge" style="background:#E24B4A;color:#fff;border-radius:20px;padding:5px 14px;font-size:12px;font-weight:700"></span>
        <button onclick="exportPaymentExcel()"
          style="background:#217346;color:#fff;border:none;border-radius:8px;padding:7px 14px;
                 font-size:12px;font-weight:600;cursor:pointer;display:flex;align-items:center;
                 gap:5px;white-space:nowrap;transition:background .15s"
          onmouseover="this.style.background='#1a5c38'" onmouseout="this.style.background='#217346'">
          📊 Export Excel
        </button>
      </div>
    </div>
    <div style="background:#fffbeb;border-bottom:1px solid #fde68a;padding:8px 20px;font-size:11px;color:#92400e;display:flex;align-items:center;gap:6px">
      <span>💡</span>
      <span>The Excel export includes <strong>Booking Created By</strong> and <strong>Remarks</strong> columns (highlighted in yellow) for manual follow-up tracking.</span>
    </div>
    <div style="overflow-x:auto;-webkit-overflow-scrolling:touch">
    <table style="width:100%;border-collapse:collapse;min-width:650px">
      <thead>
        <tr style="background:#f8fafc">
          <th style="padding:10px 14px;text-align:left;font-size:10px;font-weight:700;text-transform:uppercase;letter-spacing:.07em;color:#6b7280;border-bottom:2px solid #e2e8f0;white-space:nowrap">Room</th>
          <th style="padding:10px 14px;text-align:left;font-size:10px;font-weight:700;text-transform:uppercase;letter-spacing:.07em;color:#6b7280;border-bottom:2px solid #e2e8f0">Guest</th>
          <th style="padding:10px 14px;text-align:left;font-size:10px;font-weight:700;text-transform:uppercase;letter-spacing:.07em;color:#6b7280;border-bottom:2px solid #e2e8f0;white-space:nowrap">Conf No.</th>
          <th style="padding:10px 14px;text-align:left;font-size:10px;font-weight:700;text-transform:uppercase;letter-spacing:.07em;color:#6b7280;border-bottom:2px solid #e2e8f0">Travel Agent</th>
          <th style="padding:10px 14px;text-align:left;font-size:10px;font-weight:700;text-transform:uppercase;letter-spacing:.07em;color:#6b7280;border-bottom:2px solid #e2e8f0;white-space:nowrap">Stay</th>
          <th style="padding:10px 14px;text-align:left;font-size:10px;font-weight:700;text-transform:uppercase;letter-spacing:.07em;color:#6b7280;border-bottom:2px solid #e2e8f0;white-space:nowrap">Rate Code</th>
          <th style="padding:10px 14px;text-align:left;font-size:10px;font-weight:700;text-transform:uppercase;letter-spacing:.07em;color:#6b7280;border-bottom:2px solid #e2e8f0;white-space:nowrap">Deposit</th>
          <th style="padding:10px 14px;text-align:left;font-size:10px;font-weight:700;text-transform:uppercase;letter-spacing:.07em;color:#6b7280;border-bottom:2px solid #e2e8f0;white-space:nowrap">Action</th>
        </tr>
      </thead>
      <tbody id="pay-tbody"></tbody>
    </table>
    <div id="pay-no-results" style="text-align:center;padding:48px;color:#94a3b8;font-size:14px;display:none">
      <div style="font-size:32px;margin-bottom:10px">✅</div>
      <div style="font-weight:600;color:#374151">No outstanding payments</div>
      <div style="font-size:12px;margin-top:4px">All bookings have received deposits or are on credit accounts</div>
    </div>
    </div>
  </div>
</div>

<button class="float-back" id="float-back" onclick="showSection('summary')">
  &#8592; Back to Summary
</button>

<script>
const GUESTS = {guests_js};
const PAY_MISSING = {pay_missing_js};
// Map anchor → hl_lines dict for flag-line flashing
const HLMAP = {{}};
GUESTS.forEach(g => {{ HLMAP[g.anchor] = g.hl_lines || {{}}; }});

const FLAG_COLORS = {{
  'Platinum':'#EEEDFE:#26215C','Titanium':'#EEEDFE:#26215C','Gold':'#EEEDFE:#26215C',
  'Silver':'#EEEDFE:#26215C','Red':'#EEEDFE:#26215C',
  'VIP':'#EEEDFE:#26215C','VIP2':'#EEEDFE:#26215C','VIPR':'#EEEDFE:#26215C','VIPG':'#EEEDFE:#26215C',
  'Complaint':'#FCEBEB:#501313','Glitch':'#FCEBEB:#501313',
  'Collect':'#FAEEDA:#633806','Payment':'#FAEEDA:#633806',
  'Multi-leg':'#FAEEDA:#633806',
  'D$':'#E1F5EE:#04342C',
  'Honeymoon':'#EEEDFE:#26215C','Anniversary':'#EEEDFE:#26215C',
  'Birthday':'#EEEDFE:#26215C','Wedding':'#EEEDFE:#26215C','Babymoon':'#EEEDFE:#26215C',
  'RPT':'#E1F5EE:#04342C','Repeat':'#E1F5EE:#04342C','time RPT':'#E1F5EE:#04342C',
  'Early Check-in':'#E1F5EE:#04342C','Late Check-out':'#FAEEDA:#633806',
  'Upgrade':'#E6F1FB:#042C53','Comp':'#FBEAF0:#72243E','Child':'#FAEEDA:#633806',
  'No Flight Info':'#FCEBEB:#501313',
  'Long Stay':'#E1F5EE:#04342C',
  'Potential VIP':'#FAEEDA:#633806',
  'Restaurant':'#E1F5EE:#04342C',
  'Spa Booking':'#FBEAF0:#72243E',
  'Activities':'#E6F1FB:#042C53',
  'Sharer Missing':'#FCEBEB:#501313',
  'POA':'#FAEEDA:#633806',
  'Potential LQA':'#FCEBEB:#501313',
  'Payment Not Received':'#FCEBEB:#501313',
}};

function flagColor(f) {{
  for (const [k,v] of Object.entries(FLAG_COLORS)) {{
    if (f.includes(k)) {{ const p=v.split(':'); return `background:${{p[0]}};color:${{p[1]}}`;}}
  }}
  if (f.toLowerCase().includes('together')) return 'background:#E6F1FB;color:#042C53';
  if (/allerg|shellfish|gluten|peanut|lactose|vegan|halal/i.test(f)) return 'background:#FCEBEB;color:#501313';
  return 'background:#F1EFE8;color:#2C2C2A';
}}

let currentFilter = 'all';
let rowFlashIntervals = [];

function showSection(which) {{
  const sum = document.getElementById('section-summary');
  const rep = document.getElementById('section-report');
  const pay = document.getElementById('section-payment');
  const btn = document.getElementById('float-back');
  sum.style.display = which==='summary' ? 'block' : 'none';
  rep.style.display = which==='report'  ? 'block' : 'none';
  pay.style.display = which==='payment' ? 'block' : 'none';
  btn.classList.toggle('visible', which==='report');
  if (which !== 'report') {{
    document.querySelectorAll('.booking-flash').forEach(el => el.classList.remove('booking-flash'));
    clearHlFlash();
  }}
  ['summary','report','payment'].forEach(s => {{
    const tab = document.getElementById(`tab-${{s}}`);
    if (tab) tab.classList.toggle('active', s===which);
  }});
  if (which === 'payment') renderPayment();
}}

function exportPaymentExcel() {{
  if (!PAY_MISSING || PAY_MISSING.length === 0) {{
    alert('No outstanding payments to export.');
    return;
  }}

  // Load ExcelJS — supports full cell styling unlike SheetJS free
  const script = document.createElement('script');
  script.src = 'https://cdnjs.cloudflare.com/ajax/libs/exceljs/4.3.0/exceljs.min.js';
  script.onload = async function() {{
    const wb = new ExcelJS.Workbook();
    wb.creator = 'Arrival Highlighter';
    const ws = wb.addWorksheet('Payment Missing');

    const prop = document.querySelector('.property')?.textContent?.trim() || '';
    const arrDate = document.querySelector('.header-sub')?.textContent?.split('·')[0]?.trim() || '';
    const today = new Date().toLocaleDateString('en-GB',{{day:'2-digit',month:'short',year:'numeric'}});

    // ── Helpers ──────────────────────────────────────────────────
    const navyFill  = {{ type:'pattern', pattern:'solid', fgColor:{{argb:'FF1A1A2E'}} }};
    const amberFill = {{ type:'pattern', pattern:'solid', fgColor:{{argb:'FFF59E0B'}} }};
    const yellowFill= {{ type:'pattern', pattern:'solid', fgColor:{{argb:'FFFFFDE7'}} }};
    const evenFill  = {{ type:'pattern', pattern:'solid', fgColor:{{argb:'FFFFFFFF'}} }};
    const oddFill   = {{ type:'pattern', pattern:'solid', fgColor:{{argb:'FFF8FAFC'}} }};
    const thinBorder = {{
      top:   {{style:'thin', color:{{argb:'FFE2E8F0'}}}},
      left:  {{style:'thin', color:{{argb:'FFE2E8F0'}}}},
      bottom:{{style:'thin', color:{{argb:'FFE2E8F0'}}}},
      right: {{style:'thin', color:{{argb:'FFE2E8F0'}}}},
    }};

    // ── Column definitions ───────────────────────────────────────
    ws.columns = [
      {{key:'room',   width:9}},
      {{key:'name',   width:27}},
      {{key:'conf',   width:15}},
      {{key:'ta',     width:25}},
      {{key:'ci',     width:12}},
      {{key:'co',     width:12}},
      {{key:'nts',    width:8}},
      {{key:'rate',   width:15}},
      {{key:'dep',    width:15}},
      {{key:'createdby', width:23}},
      {{key:'remarks',   width:33}},
    ];

    // ── Row 1: Title ─────────────────────────────────────────────
    const r1 = ws.addRow(['🚨  PAYMENT NOT RECEIVED — FOLLOW UP REQUIRED','','','','','','','','','','']);
    ws.mergeCells('A1:K1');
    r1.height = 28;
    r1.getCell(1).font      = {{name:'Arial', bold:true, size:13, color:{{argb:'FFFFFFFF'}}}};
    r1.getCell(1).fill      = navyFill;
    r1.getCell(1).alignment = {{vertical:'middle', horizontal:'left'}};

    // ── Row 2: Sub-title ─────────────────────────────────────────
    const r2 = ws.addRow([`${{prop}}   |   Arrival Date: ${{arrDate}}   |   ${{PAY_MISSING.length}} booking(s) outstanding`,'','','','','','','','','','']);
    ws.mergeCells('A2:K2');
    r2.height = 20;
    r2.getCell(1).font      = {{name:'Arial', size:10, color:{{argb:'FF94A3B8'}}}};
    r2.getCell(1).fill      = navyFill;
    r2.getCell(1).alignment = {{vertical:'middle', horizontal:'left'}};

    // ── Row 3: Spacer ─────────────────────────────────────────────
    ws.addRow([]);
    ws.getRow(3).height = 6;

    // ── Row 4: Column headers ─────────────────────────────────────
    const HEADERS = ['Room','Guest Name','Conf No.','Travel Agent','Check-in','Check-out','Nights','Rate Code','Deposit (USD)','Booking Created By','Remarks'];
    const r4 = ws.addRow(HEADERS);
    r4.height = 22;
    r4.eachCell((cell, colNum) => {{
      const isManual = colNum >= 10;
      cell.font      = {{name:'Arial', bold:true, size:10,
                         color:{{argb: isManual ? 'FF92400E' : 'FFFFFFFF'}}}};
      cell.fill      = isManual ? amberFill : navyFill;
      cell.alignment = {{horizontal:'center', vertical:'middle', wrapText:true}};
      cell.border    = thinBorder;
    }});

    // ── Data rows ─────────────────────────────────────────────────
    PAY_MISSING.forEach((g, idx) => {{
      const isEven = idx % 2 === 0;
      const rowFill = isEven ? evenFill : oddFill;
      const row = ws.addRow([
        g.room, g.name, g.conf, g.ta || '',
        g.checkin || '', g.checkout || '',
        g.nights || 0, g.rate_code || '',
        g.deposit != null ? parseFloat(g.deposit.toFixed(2)) : 0,
        '', ''
      ]);
      row.height = 18;
      row.eachCell({{includeEmpty:true}}, (cell, colNum) => {{
        const isManual = colNum >= 10;
        cell.border    = thinBorder;
        cell.fill      = isManual ? yellowFill : rowFill;
        if (colNum === 1) {{
          cell.font      = {{name:'Arial', bold:true, size:11, color:{{argb:'FF1A1A2E'}}}};
          cell.alignment = {{horizontal:'center', vertical:'middle'}};
        }} else if (colNum === 9) {{
          const depVal = g.deposit || 0;
          cell.font         = {{name:'Arial', bold:true, size:10,
                               color:{{argb: depVal === 0 ? 'FFDC2626' : 'FF059669'}}}};
          cell.alignment    = {{horizontal:'right', vertical:'middle'}};
          cell.numFmt       = '#,##0.00';
        }} else if (colNum === 7) {{
          cell.font      = {{name:'Arial', size:10, color:{{argb:'FF374151'}}}};
          cell.alignment = {{horizontal:'center', vertical:'middle'}};
        }} else if (isManual) {{
          cell.font      = {{name:'Arial', size:10, italic:true, color:{{argb:'FF92400E'}}}};
          cell.alignment = {{horizontal:'left', vertical:'middle'}};
        }} else {{
          cell.font      = {{name:'Arial', size:10, color:{{argb:'FF1E2535'}}}};
          cell.alignment = {{horizontal:'left', vertical:'middle'}};
        }}
      }});
    }});

    // ── Legend row ────────────────────────────────────────────────
    ws.addRow([]);
    const legendRow = ws.addRow(['💡  Columns highlighted in amber (Booking Created By / Remarks) are for manual entry by your Reservations team.','','','','','','','','','','']);
    ws.mergeCells(`A${{legendRow.number}}:K${{legendRow.number}}`);
    legendRow.height = 18;
    legendRow.getCell(1).font      = {{name:'Arial', size:9, italic:true, color:{{argb:'FF92400E'}}}};
    legendRow.getCell(1).fill      = yellowFill;
    legendRow.getCell(1).alignment = {{horizontal:'left', vertical:'middle'}};

    // ── Freeze header row ─────────────────────────────────────────
    ws.views = [{{ state:'frozen', xSplit:0, ySplit:4, activeCell:'A5' }}];

    // ── Auto-filter ───────────────────────────────────────────────
    ws.autoFilter = {{ from:'A4', to:`K${{4 + PAY_MISSING.length}}` }};

    // ── Download ──────────────────────────────────────────────────
    const buf  = await wb.xlsx.writeBuffer();
    const blob = new Blob([buf], {{type:'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'}});
    const url  = URL.createObjectURL(blob);
    const a    = document.createElement('a');
    const dateStr = new Date().toISOString().slice(0,10).replace(/-/g,'');
    a.href     = url;
    a.download = 'PaymentMissing_' + dateStr + '.xlsx';
    document.body.appendChild(a);
    a.click();
    document.body.removeChild(a);
    URL.revokeObjectURL(url);
  }};
  script.onerror = function() {{
    alert('Could not load Excel library. Please check your internet connection.');
  }};
  document.head.appendChild(script);
}}

function renderPayment() {{
  const tbody = document.getElementById('pay-tbody');
  const noRes = document.getElementById('pay-no-results');
  const badge = document.getElementById('pay-total-badge');
  tbody.innerHTML = '';
  if (!PAY_MISSING || PAY_MISSING.length === 0) {{
    noRes.style.display = 'block';
    if (badge) badge.textContent = '0 bookings';
    return;
  }}
  noRes.style.display = 'none';
  if (badge) badge.textContent = PAY_MISSING.length + ' booking' + (PAY_MISSING.length!==1?'s':'') + ' outstanding';
  PAY_MISSING.forEach((g, idx) => {{
    const tr = document.createElement('tr');
    tr.style.background = idx%2===0 ? '#fff' : '#fafbfd';
    tr.innerHTML = `
      <td style="padding:9px 12px;font-weight:700;font-size:14px;color:#1a1a2e;vertical-align:top">${{g.room}}</td>
      <td style="padding:9px 12px;vertical-align:top">
        <div style="font-weight:600;color:#1e2535">${{g.name}}</div>
      </td>
      <td style="padding:9px 12px;font-family:monospace;font-size:11px;color:#6b7280;vertical-align:top">${{g.conf}}</td>
      <td style="padding:9px 12px;font-size:12px;color:#374151;vertical-align:top">${{g.ta || '—'}}</td>
      <td style="padding:9px 12px;font-size:11px;color:#374151;vertical-align:top;white-space:nowrap">
        ${{g.checkin}} → ${{g.checkout}}<br>
        <span style="color:#6b7280">${{g.nights}}N</span>
      </td>
      <td style="padding:9px 12px;font-family:monospace;font-size:11px;color:#6b7280;vertical-align:top">${{g.rate_code}}</td>
      <td style="padding:9px 12px;vertical-align:top">
        <span style="background:#FCEBEB;color:#501313;font-size:11px;padding:2px 8px;border-radius:8px;font-weight:600">USD ${{g.deposit.toFixed(2)}}</span>
      </td>
      <td style="padding:9px 12px;vertical-align:top">
        <button onclick="goToBooking('${{g.anchor}}',null,null)"
          style="font-size:11px;background:#1a56db;color:#fff;border:none;border-radius:8px;
                 padding:5px 12px;cursor:pointer;font-weight:600;white-space:nowrap">
          📄 View in Report
        </button>
      </td>`;
    tbody.appendChild(tr);
  }});
}}

function saveNote(el) {{
  if (el.value.trim()) {{
    localStorage.setItem(el.dataset.key, el.value);
  }} else {{
    localStorage.removeItem(el.dataset.key);
  }}
}}

// ── Text highlight in notes ───────────────────────────────────────────
const _hlToolbar = document.createElement('div');
_hlToolbar.className = 'note-hl-toolbar';
_hlToolbar.style.display = 'none';
_hlToolbar.innerHTML = `
  <button onclick="_applyNoteHl()" title="Highlight yellow">🟡 Highlight</button>
  <button onclick="_removeNoteHl()" title="Remove highlight">✕</button>`;
document.body.appendChild(_hlToolbar);

let _hlActiveEl = null;

document.addEventListener('mouseup', function(e) {{
  // Ignore clicks inside toolbar itself
  if (_hlToolbar.contains(e.target)) return;
  const noteEl = e.target.closest('.note-original-text');
  if (!noteEl) {{ _hlToolbar.style.display = 'none'; return; }}
  const sel = window.getSelection();
  if (!sel || sel.isCollapsed || !sel.toString().trim()) {{
    _hlToolbar.style.display = 'none'; return;
  }}
  _hlActiveEl = noteEl;
  // Position toolbar near selection
  const rect = sel.getRangeAt(0).getBoundingClientRect();
  const scrollY = window.scrollY || document.documentElement.scrollTop;
  _hlToolbar.style.display = 'flex';
  _hlToolbar.style.left = Math.max(4, rect.left + rect.width/2 - 70) + 'px';
  _hlToolbar.style.top  = (rect.top + scrollY - 42) + 'px';
  _hlToolbar.style.position = 'absolute';
}});

function _applyNoteHl() {{
  const sel = window.getSelection();
  if (!sel || sel.isCollapsed || !_hlActiveEl) {{ _hlToolbar.style.display='none'; return; }}
  const range = sel.getRangeAt(0);
  // Only operate within the note element
  if (!_hlActiveEl.contains(range.commonAncestorContainer)) {{ _hlToolbar.style.display='none'; return; }}
  const span = document.createElement('span');
  span.className = 'note-hl';
  try {{
    range.surroundContents(span);
  }} catch(e) {{
    // Selection crosses element boundaries — extract and wrap
    const frag = range.extractContents();
    span.appendChild(frag);
    range.insertNode(span);
  }}
  sel.removeAllRanges();
  _hlToolbar.style.display = 'none';
  _saveNoteHls(_hlActiveEl);
}}

function _removeNoteHl() {{
  const sel = window.getSelection();
  if (!sel || sel.isCollapsed) {{ _hlToolbar.style.display='none'; return; }}
  const range = sel.getRangeAt(0);
  // Unwrap any note-hl spans within the selection
  const container = range.commonAncestorContainer.nodeType === 3
    ? range.commonAncestorContainer.parentElement
    : range.commonAncestorContainer;
  const hlSpans = container.querySelectorAll ? container.querySelectorAll('.note-hl') : [];
  hlSpans.forEach(sp => {{
    if (sel.containsNode(sp, true)) {{
      const parent = sp.parentNode;
      while (sp.firstChild) parent.insertBefore(sp.firstChild, sp);
      parent.removeChild(sp);
    }}
  }});
  // Also check if we're inside a hl span
  let el = range.commonAncestorContainer;
  while (el && el !== document.body) {{
    if (el.classList && el.classList.contains('note-hl')) {{
      const parent = el.parentNode;
      while (el.firstChild) parent.insertBefore(el.firstChild, el);
      parent.removeChild(el);
      break;
    }}
    el = el.parentNode;
  }}
  sel.removeAllRanges();
  _hlToolbar.style.display = 'none';
  if (_hlActiveEl) _saveNoteHls(_hlActiveEl);
}}

function _saveNoteHls(noteEl) {{
  // Save highlights as the innerHTML so they persist in localStorage
  const row = noteEl.closest('tr');
  if (!row) return;
  const textarea = row.querySelector('.note-manual-field');
  if (!textarea) return;
  const key = 'notehl-' + textarea.dataset.key;
  localStorage.setItem(key, noteEl.innerHTML);
}}

function _loadNoteHls(noteEl, key) {{
  const saved = localStorage.getItem('notehl-' + key);
  if (saved) noteEl.innerHTML = saved;
}}

function styleNote(el, original) {{}} // no longer needed

function downloadPDF() {{
  const style = document.createElement('style');
  style.id = 'print-style';
  style.textContent = `
    @media print {{
      #section-report {{ display: none !important; }}
      #section-payment {{ display: none !important; }}
      #section-summary {{ display: block !important; zoom: 100% !important; }}
      .table-wrap {{ padding: 0; }}
      table {{ font-size: 10px; }}
      .note-manual-field {{ border: 1px dashed #fca5a5; resize: none; background: #fff5f5; }}
      @page {{ size: landscape; margin: 10mm; }}
    }}`;
  document.head.appendChild(style);
  window.print();
  setTimeout(() => {{ const s = document.getElementById('print-style'); if(s) s.remove(); }}, 1000);
}}

let activeHlIds = [];
let hlFlashInterval = null;

function clearHlFlash() {{
  if (hlFlashInterval) {{ clearInterval(hlFlashInterval); hlFlashInterval = null; }}
  activeHlIds.forEach(id => {{
    const el = document.querySelector(`[data-hlid="${{id}}"]`);
    if (el) {{
      el.classList.remove('hl-line-flash');
      el.style.background = '';
      el.style.boxShadow = '';
      el.style.zIndex = '';
    }}
  }});
  activeHlIds = [];
}}

function flashHlLines(hlIds) {{
  clearHlFlash();
  if (!hlIds || !hlIds.length) return;
  activeHlIds = hlIds;
  activeHlIds.forEach(id => {{
    const el = document.querySelector(`[data-hlid="${{id}}"]`);
    if (el) {{
      el.classList.add('hl-line-flash');
    }}
  }});
}}

function goToBooking(anchor, flagCat, hlLines) {{
  showSection('report');
  clearHlFlash();
  setTimeout(() => {{
    const el = document.getElementById(anchor);
    if (el) {{
      el.scrollIntoView({{behavior:'smooth', block:'start'}});
      el.classList.remove('booking-flash');
      void el.offsetWidth;
      el.classList.add('booking-flash');
    }}
    let idsToFlash = [];
    if (hlLines) {{
      if (hlLines['booking']) idsToFlash = [...hlLines['booking']];
      if (flagCat && hlLines[flagCat]) {{
        hlLines[flagCat].forEach(id => {{ if (!idsToFlash.includes(id)) idsToFlash.push(id); }});
      }}
    }}
    if (idsToFlash.length) {{
      flashHlLines(idsToFlash);
      setTimeout(() => {{
        const catIds = (flagCat && hlLines && hlLines[flagCat]) ? hlLines[flagCat] : idsToFlash;
        const targetId = catIds.find(id => !id.startsWith('anchor-')) || catIds[0];
        const targetEl = document.querySelector(`[data-hlid="${{targetId}}"]`);
        if (targetEl) targetEl.scrollIntoView({{behavior:'smooth', block:'center'}});
      }}, 500);
    }}
  }}, 100);
}}

function render(filterCat) {{
  currentFilter = filterCat;
  const tbody = document.getElementById('tbody');
  tbody.innerHTML = '';
  rowFlashIntervals.forEach(id => clearInterval(id));
  rowFlashIntervals = [];
  let shown = 0;
  // Build set of excluded GHA levels from checkboxes
  const excludedLevels = new Set();
  document.querySelectorAll('.gha-level-toggle').forEach(cb => {{
    if (!cb.checked) excludedLevels.add(cb.dataset.level);
  }});
  GUESTS.forEach(g => {{
    // Detect GHA level — exact match on lowercase flag text
    let ghaLevel = null;
    for (const f of g.flags) {{
      const fl = f.toLowerCase().trim();
      if (fl === 'silver')   {{ ghaLevel = 'silver';   break; }}
      if (fl === 'gold')     {{ ghaLevel = 'gold';     break; }}
      if (fl === 'red')      {{ ghaLevel = 'red';      break; }}
      if (fl === 'platinum') {{ ghaLevel = 'platinum'; break; }}
      if (fl === 'titanium') {{ ghaLevel = 'titanium'; break; }}
    }}
    // Skip entirely when level is unchecked and GHA filter is active
    if (ghaLevel && excludedLevels.has(ghaLevel) && filterCat === 'membership') return;
    const match = (filterCat === 'all' || g.cats.includes(filterCat));
    const tr = document.createElement('tr');
    tr.dataset.cats = g.cats.join(' ');
    if (!match) tr.classList.add('dimmed');

    const flightHtml = g.flight && g.flight !== '--' && g.flight !== '-' && g.flight !== ''
      ? (() => {{
          if (g.flight === 'NO FLIGHT INFO') {{
            return `<span class="flight-badge noeta">❓ No Flight Info</span>`;
          }}
          const isTransport = /^(RMV|OTH|Airport|FERRY|TRF)$/i.test(g.flight.trim());
          const noEta = !isTransport && g.flight.includes('NO ETA');
          const parts = g.flight.replace(' NO ETA','').trim().split(' ');
          const flightNo = parts[0];
          const eta = parts.length > 1 && !noEta && !isTransport ? parts[1] : '';
          const method = g.arr_method || '';
          let badge = '';
          if (isTransport) {{
            badge = `<span class="flight-badge" style="background:#f1f5f9;color:#475569">${{flightNo}}</span>`;
          }} else if (noEta) {{
            badge = `<span class="flight-badge noeta">${{flightNo}} <span style="font-size:10px;opacity:.85">⚠️ No ETA</span></span>`;
          }} else if (eta) {{
            badge = `<span class="flight-badge">${{flightNo}} <span style="font-size:10px;color:#166534;background:#dcfce7;padding:1px 5px;border-radius:8px;margin-left:2px">🕐 ${{eta}}</span></span>`;
          }} else {{
            badge = `<span class="flight-badge">${{flightNo}}</span>`;
          }}
          if (method) badge += `<br><span style="font-size:10px;color:#6b7280;margin-top:2px;display:inline-block;letter-spacing:.04em">${{method}}</span>`;
          return badge;
        }})()
      : '<span style="color:#cbd5e1">—</span>';

    const flagsHtml = g.flags.map(f => {{
      const fl = f.toLowerCase();
      let cat = 'roomno';
      if (/platinum|titanium|gold|silver|red/.test(fl)) cat='membership';
      else if (/vip/.test(fl)) cat='vip';
      else if (/rpt|repeat|time rpt/.test(fl)) cat='repeater';
      else if (/allerg|shellfish|gluten|peanut|lactose|vegan|halal/.test(fl)) cat='allergy';
      else if (/complaint|glitch/.test(fl)) cat='complaint';
      else if (/collect|payment/.test(fl)) cat='payment';
      else if (/honeymoon|anniversary|birthday|wedding|babymoon|proposal/.test(fl)) cat='occasion';
      else if (/together/.test(fl)) cat='together';
      else if (/multi-leg|leg/.test(fl)) cat='legs';
      else if (fl.includes('d$')) cat='dbalance';
      else if (/early check-in/.test(fl)) cat='eci';
      else if (/late check-out/.test(fl)) cat='lco';
      else if (/upgrade/.test(fl)) cat='upgrade';
      else if (/^comp$/.test(fl)) cat='comp';
      else if (/child/.test(fl)) cat='children';
      else if (/no flight info/.test(fl)) cat='noflight';
      else if (/long stay/.test(fl)) cat='longstay';
      else if (/potential vip/.test(fl)) cat='potvip';
      else if (/restaurant/.test(fl)) cat='restaurant';
      else if (/spa booking/.test(fl)) cat='spa';
      else if (/activit/.test(fl)) cat='activity';
      else if (/sharer missing/.test(fl)) cat='sharermissing';
      else if (/^poa/.test(fl)) cat='poa';
      else if (/potential lqa/.test(fl)) cat='lqa';
      else if (/payment not received/.test(fl)) cat='paymissing';
      else if (/ek|ey|qr|g9|ku|gf|sq|ai|6e/.test(fl)) cat='flight';
      return `<span class="flag" style="${{flagColor(f)}}" title="Jump to related lines in report"
        onclick="goToBooking('${{g.anchor}}','${{cat}}',HLMAP['${{g.anchor}}'])">${{f}}</span>`;
    }}).join('');

    tr.innerHTML = `
      <td><span class="room-no">${{g.room}}</span><br><span class="conf-no">${{g.conf}}</span>${{g.room_type ? `<br><span style="font-size:10px;color:#374151;font-family:monospace;font-weight:600">${{g.room_type}}</span>` : ''}}${{g.rate_code ? `<br><span style="font-size:9px;color:#6b7280;font-family:monospace">${{g.rate_code}}</span>` : ''}}</td>
      <td>
        <span class="guest-name guest-link" onclick="goToBooking('${{g.anchor}}',null,null)" title="Jump to booking in report">${{g.name}}</span>
        ${{g.ta ? `<br><span class="ta-name">${{g.ta}}</span>` : ''}}
        ${{g.checkin && g.checkout ? `<br><span style="font-size:10px;color:#1e2535;font-weight:500">${{g.checkin}} → ${{g.checkout}}${{g.nights ? ` (${{g.nights}}N)` : ''}}</span>` : ''}}
        ${{(g.adults || g.children) ? `<br><span style="font-size:10px;color:#374151">👤 ${{g.adults}}${{g.children ? ` · 🧒 ${{g.children}}` : ''}}</span>` : ''}}
      </td>
      <td>${{flightHtml}}</td>
      <td><div class="flags">${{flagsHtml}}</div></td>
      <td>
        <div class="note-original-text">${{g.note}}</div>
        <textarea class="note-manual-field" data-key="note-${{g.room}}-${{g.conf}}"
          placeholder="+ Add notes..."
          oninput="saveNote(this)"
        ></textarea>
      </td>`;
    tbody.appendChild(tr);
    const noteEl = tr.querySelector('.note-manual-field');
    if (noteEl) {{
      const saved = localStorage.getItem(noteEl.dataset.key);
      if (saved) noteEl.value = saved;
    }}
    const noteTextEl = tr.querySelector('.note-original-text');
    if (noteTextEl) {{
      _loadNoteHls(noteTextEl, `note-${{g.room}}-${{g.conf}}`);
    }}

    if (match) {{
      shown++;
      if (filterCat !== 'all') {{
        let flashOn = true;
        tr.style.background = '#FFFDE7';
        const iv = setInterval(() => {{
          if (currentFilter !== filterCat) {{
            tr.style.background = '';
            clearInterval(iv);
            return;
          }}
          flashOn = !flashOn;
          tr.style.background = flashOn ? '#FFFDE7' : '#fff';
        }}, 900);
        rowFlashIntervals.push(iv);
      }}
    }}
  }});
  document.getElementById('no-results').style.display = shown === 0 ? 'block' : 'none';
  document.getElementById('count-badge').textContent = shown + ' guest' + (shown!==1?'s':'');
}}

function filterGuests(cat) {{
  if (cat === 'paymissing') {{ showSection('payment'); return; }}
  if (currentFilter === cat) cat = 'all';
  rowFlashIntervals.forEach(id => clearInterval(id));
  rowFlashIntervals = [];
  document.querySelectorAll('#tbody tr').forEach(tr => tr.style.background = '');
  document.querySelectorAll('.stat-card').forEach(c => c.classList.remove('active'));
  const card = document.querySelector(`[data-cat="${{cat}}"]`);
  if (card) card.classList.add('active');
  const labels = {{
    'all':'All Bookings','repeater':'Repeater Guests','flight':'Flights',
    'membership':'GHA Members','vip':'VIP Guests','complaint':'Complaints / Glitches',
    'allergy':'Allergies / Dietary','occasion':'Special Occasions','payment':'Payments to Collect',
    'dbalance':'D$ Balance','together':'Travelling Together','legs':'Multi-Villa Legs','noeta':'Missing ETA',
    'eci':'Early Check-in','lco':'Late Check-out','upgrade':'Upgrades','comp':'Complimentary','children':'With Children',
    'noflight':'No Flight Info',
    'longstay':'Long Stay (10+ nights)',
    'potvip':'Potential VIP (title in notes)',
    'restaurant':'Restaurant Bookings',
    'spa':'Spa Bookings',
    'activity':'Activities',
    'sharermissing':'Sharer Name Missing',
    'poa':'POA with Amount',
    'lqa':'Potential LQA Auditor',
    'paymissing':'Payment Not Received'
  }};
  document.getElementById('filter-tag').textContent = labels[cat] || cat;
  updateGHACount();
  render(cat);
}}

function updateGHACount() {{
  const excludedLevels = new Set();
  document.querySelectorAll('.gha-level-toggle').forEach(cb => {{
    if (!cb.checked) excludedLevels.add(cb.dataset.level);
  }});
  const ghaCard = document.querySelector('[data-cat="membership"] .stat-num');
  if (!ghaCard) return;
  const count = GUESTS.filter(g => {{
    if (!g.cats.includes('membership')) return false;
    let ghaLevel = null;
    for (const f of g.flags) {{
      const fl = f.toLowerCase().trim();
      if (fl === 'silver')   {{ ghaLevel = 'silver';   break; }}
      if (fl === 'gold')     {{ ghaLevel = 'gold';     break; }}
      if (fl === 'red')      {{ ghaLevel = 'red';      break; }}
      if (fl === 'platinum') {{ ghaLevel = 'platinum'; break; }}
      if (fl === 'titanium') {{ ghaLevel = 'titanium'; break; }}
    }}
    if (ghaLevel && excludedLevels.has(ghaLevel)) return false;
    return true;
  }}).length;
  ghaCard.textContent = count;
}}

render('all');
</script>
</body>
</html>'''
    return html





def build_briefing_html(summary_data, pdf_filename='highlighted.pdf'):
    """Priority Briefing Sheet — only guests needing attention, VIP-first ordering."""
    import json, html as htmllib

    guests  = summary_data['guests']
    prop    = summary_data.get('property', '')
    dt      = summary_data.get('date', '')

    # ── Priority tiers ────────────────────────────────────────────────────
    TIER1_FLAGS = {
        'VIP','VIP2','VIPG','VIPR','Platinum','Titanium','Gold','Silver','Red',
        'Potential VIP','Potential LQA',
    }
    TIER2_FLAGS = {
        'Complaint','Allergy','Allergic','Collect','No Flight Info',
        'Sharer Missing','POA',
    }
    TIER3_FLAGS = {
        'Honeymoon','Wedding','Birthday','Anniversary','Babymoon',
        'Proposal','Engagement','Birthday Cake','Birthday Amenity',
        'Birthday Decoration','Honeymoon Setup','Honeymoon Amenity',
        'Restaurant Booking','Spa Booking','Activities','Upgrade','Comp',
        'Children',
    }
    WORTH_NOTING = {
        'RPT','Repeat','Early Check-in','Together','Multi-leg',
    }

    def get_tier(flags):
        fl_lower = [f.lower() for f in flags]
        for f in flags:
            if any(f.startswith(t) or f == t for t in TIER1_FLAGS): return 1
            if any(t.lower() in f.lower() for t in {'Potential VIP','Potential LQA',
               'VIP2','VIPG','VIPR','Platinum','Titanium','Gold'}): return 1
        for f in flags:
            if any(t.lower() in f.lower() for t in TIER2_FLAGS): return 2
        for f in flags:
            if any(t.lower() in f.lower() for t in TIER3_FLAGS): return 3
        for f in flags:
            if any(t.lower() in f.lower() for t in WORTH_NOTING): return 4
        return 99  # exclude

    def flag_badge(f):
        fl = f.lower()
        if any(x in fl for x in ['platinum','titanium','gold','silver','red','vip2','vipg','vipr']):
            bg,fg = '#EEEDFE','#26215C'
        elif 'potential vip' in fl or 'potential lqa' in fl:
            bg,fg = '#FCEBEB','#501313'
        elif any(x in fl for x in ['allerg','shellfish','gluten','peanut','lactose','vegan','halal']):
            bg,fg = '#FCEBEB','#501313'
        elif any(x in fl for x in ['complaint','glitch']):
            bg,fg = '#FCEBEB','#501313'
        elif any(x in fl for x in ['collect','payment','poa']):
            bg,fg = '#FAEEDA','#633806'
        elif any(x in fl for x in ['honeymoon','wedding','birthday','anniversary',
                                    'babymoon','proposal','engagement']):
            bg,fg = '#EEEDFE','#26215C'
        elif 'no flight' in fl or 'sharer missing' in fl:
            bg,fg = '#FCEBEB','#501313'
        elif any(x in fl for x in ['restaurant','spa booking']):
            bg,fg = '#E1F5EE','#04342C'
        elif 'activit' in fl:
            bg,fg = '#E6F1FB','#042C53'
        elif any(x in fl for x in ['rpt','repeat','stays']):
            bg,fg = '#E1F5EE','#04342C'
        elif any(x in fl for x in ['early check-in','upgrade','comp','together','multi-leg']):
            bg,fg = '#FAEEDA','#633806'
        elif any(x in fl for x in ['long stay','child']):
            bg,fg = '#E1F5EE','#04342C'
        else:
            bg,fg = '#F1EFE8','#2C2C2A'
        label = htmllib.escape(f)
        return f'<span style="display:inline-block;font-size:10px;padding:2px 8px;border-radius:10px;font-weight:600;background:{bg};color:{fg};margin:2px 2px 2px 0">{label}</span>'

    # ── Filter and sort guests ────────────────────────────────────────────
    briefing = []
    for g in guests:
        tier = get_tier(g['flags'])
        if tier <= 4:
            briefing.append((tier, g))

    briefing.sort(key=lambda x: x[1]['room'].zfill(4))

    total_rooms  = summary_data['rooms']
    briefing_cnt = len(briefing)

    # ── Tier labels ───────────────────────────────────────────────────────
    TIER_META = {
        1: ('🔴', 'VIP & High Profile',   '#FCEBEB', '#501313'),
        2: ('🟠', 'Needs Action',          '#FEF3CD', '#7B4F00'),
        3: ('🟡', 'Needs Preparation',     '#FEFCE8', '#6B5F00'),
        4: ('🔵', 'Worth Noting',          '#EFF6FF', '#1E3A5F'),
    }

    # Build rows grouped by tier
    rows_html = ''
    for tier, g in briefing:
        icon, label, bg, fg = TIER_META[tier]

        name    = htmllib.escape(g['name'])
        ta      = htmllib.escape(g.get('ta',''))
        flight  = g.get('flight','')
        checkin = g.get('checkin','')
        checkout= g.get('checkout','')
        nights  = g.get('nights',0)
        adults  = g.get('adults',0)
        children= g.get('children',0)
        room_type = g.get('room_type','')
        rate_code = g.get('rate_code','')

        # Flight badge
        if flight == 'NO FLIGHT INFO':
            fl_badge = '<span style="background:#FCEBEB;color:#501313;font-size:10px;padding:2px 7px;border-radius:8px;font-weight:600">❓ No Flight</span>'
        elif flight and 'NO ETA' in flight:
            fl_badge = f'<span style="background:#FAEEDA;color:#633806;font-size:10px;padding:2px 7px;border-radius:8px;font-weight:600">✈️ {htmllib.escape(flight)}</span>'
        elif flight:
            fl_badge = f'<span style="background:#E1F5EE;color:#04342C;font-size:10px;padding:2px 7px;border-radius:8px;font-weight:600">✈️ {htmllib.escape(flight)}</span>'
        else:
            fl_badge = '<span style="color:#94a3b8;font-size:10px">—</span>'

        flags_html = ''.join(flag_badge(f) for f in g['flags'])

        pax = f'👤 {adults}'
        if children: pax += f' · 🧒 {children}'

        rows_html += f'''
        <tr class="guest-row">
          <td style="font-weight:700;font-size:15px;color:#1a1a2e;white-space:nowrap">
            <span style="font-size:9px;margin-right:3px">{icon}</span>{g["room"]}
            <div style="font-size:9px;color:#94a3b8;font-family:monospace;font-weight:400">{room_type}</div>
            <div style="font-size:9px;color:#b0b8c4;font-family:monospace;font-weight:400">{rate_code}</div>
          </td>
          <td>
            <div style="font-weight:600;color:#1e2535;font-size:13px">{name}</div>
            {f'<div style="font-size:10px;color:#6b7280;font-style:italic">{ta}</div>' if ta else ''}
          </td>
          <td style="white-space:nowrap">
            <div style="font-size:11px;color:#1e2535;font-weight:500">{checkin} → {checkout}</div>
            <div style="font-size:10px;color:#6b7280">{f"{nights}N" if nights else ""} &nbsp; {pax}</div>
          </td>
          <td>{fl_badge}</td>
          <td><div style="display:flex;flex-wrap:wrap;gap:2px">{flags_html}</div></td>
          <td style="font-size:10px;color:#374151;max-width:180px">{htmllib.escape(g.get("note",""))}</td>
        </tr>'''

    if not briefing:
        rows_html = '<tr><td colspan="6" style="text-align:center;padding:40px;color:#94a3b8">No priority guests for this arrival</td></tr>'

    return f'''<!DOCTYPE html>
<html lang="en">
<head>
<meta charset="UTF-8">
<meta name="viewport" content="width=device-width,initial-scale=1">
<title>Morning Briefing — {dt}</title>
<style>
*{{box-sizing:border-box;margin:0;padding:0}}
body{{font-family:-apple-system,BlinkMacSystemFont,'Segoe UI',sans-serif;background:#f1f5f9;color:#1e2535;font-size:13px}}
.header{{background:#1a1a2e;color:#fff;padding:14px 24px;display:flex;align-items:center;justify-content:space-between;position:sticky;top:0;z-index:100}}
.header h1{{font-size:17px;font-weight:700}}
.header-sub{{font-size:11px;color:#94a3b8;margin-top:2px}}
.property{{font-size:17px;font-weight:700;color:#fff;text-align:right}}
.meta-bar{{background:#fff;border-bottom:1px solid #e2e8f0;padding:10px 24px;
           display:flex;align-items:center;gap:20px;font-size:12px;color:#6b7280;flex-wrap:wrap}}
.meta-bar strong{{color:#1e2535}}
.badge{{background:#1a56db;color:#fff;border-radius:20px;padding:2px 10px;font-size:11px;font-weight:600}}
.tier-legend{{margin-left:auto;font-size:11px;display:flex;gap:12px;flex-wrap:wrap}}
.wrap{{padding:16px 24px}}
table{{width:100%;border-collapse:collapse;background:#fff;border-radius:10px;overflow:hidden;box-shadow:0 1px 3px rgba(0,0,0,.06)}}
th{{background:#f8fafc;font-size:10px;font-weight:700;text-transform:uppercase;letter-spacing:.07em;color:#6b7280;padding:9px 11px;text-align:left;border-bottom:1px solid #e2e8f0}}
td{{padding:9px 11px;border-bottom:1.5px solid #f1f5f9;vertical-align:top}}
.guest-row:nth-child(even){{background:#fafbfd}}
.guest-row:hover{{background:#f0f7ff!important}}
.tier-header td{{font-size:10px;font-weight:700;padding:5px 12px;letter-spacing:.06em;text-transform:uppercase;border-bottom:1px solid #e2e8f0}}
.footer{{text-align:center;padding:16px;font-size:10px;color:#94a3b8}}
@media print{{
  .header,.meta-bar{{-webkit-print-color-adjust:exact;print-color-adjust:exact}}
  .wrap{{padding:4px}}
  .guest-row{{page-break-inside:avoid}}
}}
@media (max-width:768px){{
  .header{{flex-wrap:wrap;gap:4px;padding:10px 14px}}
  .property{{width:100%;text-align:left;font-size:13px}}
  .meta-bar{{padding:8px 12px}}
  .tier-legend{{margin-left:0;width:100%}}
  .wrap{{padding:8px 0;overflow-x:auto}}
  table{{min-width:620px}}
}}
</style>
</head>
<body>
<div class="header">
  <div>
    <h1>📋 Morning Briefing</h1>
    <div class="header-sub">{dt} &nbsp;·&nbsp; Priority Arrivals Only</div>
  </div>
  <div class="property">{prop}</div>
</div>
<div class="meta-bar">
  <div>Total Arrivals: <strong>{total_rooms}</strong></div>
  <div>In Briefing: <strong>{briefing_cnt}</strong> &nbsp;<span class="badge">{briefing_cnt} of {total_rooms}</span></div>
  <div class="tier-legend">
    <span>🔴 VIP &amp; High Profile</span>
    <span>🟠 Needs Action</span>
    <span>🟡 Needs Preparation</span>
    <span>🔵 Worth Noting</span>
  </div>
</div>
<div class="wrap">
<table>
  <thead>
    <tr>
      <th style="width:75px">Room</th>
      <th style="width:170px">Guest</th>
      <th style="width:130px">Stay / Pax</th>
      <th style="width:110px">Flight</th>
      <th>Flags</th>
      <th style="width:180px">Notes</th>
    </tr>
  </thead>
  <tbody>
    {rows_html}
  </tbody>
</table>
</div>
<div class="footer">Morning Briefing &nbsp;·&nbsp; {dt} &nbsp;·&nbsp; {prop}</div>
</body>
</html>'''


def build_payment_excel(summary_data):
    """Build a properly styled .xlsx for Payment Not Received bookings."""
    from openpyxl import Workbook
    from openpyxl.styles import (Font, PatternFill, Alignment, Border, Side,
                                  GradientFill)
    from openpyxl.utils import get_column_letter
    import io as _io

    guests = summary_data.get('guests', [])
    pay_missing = [g for g in guests if any('Payment Not Received' in f for f in g['flags'])]

    wb = Workbook()
    ws = wb.active
    ws.title = 'Payment Missing'

    # ── Colour palette ──────────────────────────────────────────
    C_HEADER_BG  = '1A1A2E'   # dark navy
    C_HEADER_FG  = 'FFFFFF'
    C_TITLE_BG   = 'E24B4A'   # red accent
    C_COL_HDR_BG = 'F8FAFC'
    C_COL_HDR_FG = '374151'
    C_ROW_EVEN   = 'FFFFFF'
    C_ROW_ODD    = 'F8FAFC'
    C_MANUAL_BG  = 'FFFDE7'   # yellow — manual entry columns
    C_MANUAL_FG  = '92400E'
    C_BORDER     = 'E2E8F0'
    C_DEPOSIT_FG = 'DC2626'   # red for zero deposit

    def fill(hex_):  return PatternFill('solid', fgColor=hex_)
    def font(bold=False, sz=10, color='1E2535', italic=False):
        return Font(name='Arial', bold=bold, size=sz, color=color, italic=italic)
    def border():
        s = Side(style='thin', color=C_BORDER)
        return Border(left=s, right=s, top=s, bottom=s)
    def center(): return Alignment(horizontal='center', vertical='center', wrap_text=True)
    def left():   return Alignment(horizontal='left',   vertical='center', wrap_text=False)
    def right():  return Alignment(horizontal='right',  vertical='center')

    # ── Title row (row 1) ────────────────────────────────────────
    ws.merge_cells('A1:K1')
    ws['A1'] = '🚨  PAYMENT NOT RECEIVED — FOLLOW UP REQUIRED'
    ws['A1'].font      = Font(name='Arial', bold=True, size=13, color=C_HEADER_FG)
    ws['A1'].fill      = fill(C_HEADER_BG)
    ws['A1'].alignment = Alignment(horizontal='left', vertical='center')
    ws.row_dimensions[1].height = 28

    # ── Sub-title row (row 2) ────────────────────────────────────
    prop = summary_data.get('property','')
    dt   = summary_data.get('date','')
    ws.merge_cells('A2:K2')
    ws['A2'] = f"{prop}   |   Arrival Date: {dt}   |   {len(pay_missing)} booking(s) outstanding"
    ws['A2'].font      = Font(name='Arial', size=10, color='94A3B8')
    ws['A2'].fill      = fill(C_HEADER_BG)
    ws['A2'].alignment = Alignment(horizontal='left', vertical='center')
    ws.row_dimensions[2].height = 20

    # ── Blank spacer row (row 3) ─────────────────────────────────
    ws.row_dimensions[3].height = 6

    # ── Column headers (row 4) ───────────────────────────────────
    COLS = [
        ('Room',               8,  False),
        ('Guest Name',         26, False),
        ('Conf No.',           14, False),
        ('Travel Agent',       24, False),
        ('Check-in',           11, False),
        ('Check-out',          11, False),
        ('Nights',              7, False),
        ('Rate Code',          14, False),
        ('Deposit (USD)',       14, False),
        ('Booking Created By', 22, True),   # manual — yellow
        ('Remarks',            32, True),   # manual — yellow
    ]
    for ci, (label, width, is_manual) in enumerate(COLS, start=1):
        cell = ws.cell(row=4, column=ci, value=label)
        cell.font      = Font(name='Arial', bold=True, size=10,
                               color='FFFFFF' if not is_manual else C_MANUAL_FG)
        cell.fill      = fill('2D3748' if not is_manual else 'F59E0B')
        cell.alignment = center()
        cell.border    = border()
        ws.column_dimensions[get_column_letter(ci)].width = width
    ws.row_dimensions[4].height = 22

    # ── Data rows ─────────────────────────────────────────────────
    for ri, g in enumerate(pay_missing):
        row_num  = 5 + ri
        is_even  = ri % 2 == 0
        row_bg   = C_ROW_EVEN if is_even else C_ROW_ODD

        values = [
            g['room'],
            g['name'],
            g.get('conf',''),
            g.get('ta',''),
            g.get('checkin',''),
            g.get('checkout',''),
            g.get('nights', 0),
            g.get('rate_code',''),
            g.get('deposit', 0.0),
            '',   # Booking Created By
            '',   # Remarks
        ]
        for ci, val in enumerate(values, start=1):
            cell = ws.cell(row=row_num, column=ci, value=val)
            is_manual = ci >= 10
            cell.border = border()
            cell.font   = font(sz=10, color=C_MANUAL_FG if is_manual else '1E2535',
                               italic=is_manual)
            cell.fill   = fill(C_MANUAL_BG if is_manual else row_bg)
            if ci == 1:   # Room — bold centred
                cell.font      = font(bold=True, sz=11)
                cell.alignment = center()
            elif ci == 9:  # Deposit — right-aligned, red if zero
                cell.number_format = '#,##0.00'
                cell.alignment     = right()
                cell.font          = font(bold=True, sz=10,
                                          color=C_DEPOSIT_FG if (val or 0) == 0 else '059669')
            elif ci == 7:  # Nights — centred
                cell.alignment = center()
            else:
                cell.alignment = left()
        ws.row_dimensions[row_num].height = 18

    # ── No data message ───────────────────────────────────────────
    if not pay_missing:
        ws.merge_cells('A5:K5')
        ws['A5'] = '✅  No outstanding payments — all bookings have deposits or are on credit accounts'
        ws['A5'].font      = Font(name='Arial', size=11, color='059669', italic=True)
        ws['A5'].alignment = Alignment(horizontal='center', vertical='center')
        ws.row_dimensions[5].height = 30

    # ── Freeze panes below header row ────────────────────────────
    ws.freeze_panes = 'A5'

    # ── Auto-filter ───────────────────────────────────────────────
    if pay_missing:
        ws.auto_filter.ref = f"A4:K{4 + len(pay_missing)}"

    # ── Legend row at bottom ──────────────────────────────────────
    legend_row = 6 + len(pay_missing)
    ws.merge_cells(f'A{legend_row}:K{legend_row}')
    ws[f'A{legend_row}'] = '💡  Columns highlighted in amber (Booking Created By / Remarks) are for manual entry by your Reservations team.'
    ws[f'A{legend_row}'].font      = Font(name='Arial', size=9, color='92400E', italic=True)
    ws[f'A{legend_row}'].fill      = fill('FFFDE7')
    ws[f'A{legend_row}'].alignment = Alignment(horizontal='left', vertical='center')
    ws.row_dimensions[legend_row].height = 18

    buf = _io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.getvalue()


# ── UI ───────────────────────────────────────────────────────────────────
st.markdown("## 🏝️ Arrival Report Highlighter")
st.markdown("**Anantara Veli · Anantara Dhigu · Naladhu Private Island**")
st.markdown("---")

col_left, col_right = st.columns([1, 2])

with col_left:
    st.markdown("### 📂 Upload Reports")
    uploaded_files = st.file_uploader(
        "Upload one or more arrival PDFs", type=['pdf'],
        accept_multiple_files=True, help="Upload AVEL, ADHI, MNLD PDFs")
    st.markdown("### ⚙️ Categories")
    enabled = {}
    for cat_id, cat in CATEGORIES.items():
        enabled[cat_id] = st.checkbox(
            f"{cat['icon']} {cat['label']}", value=True, key=f"cat_{cat_id}")
    process_btn = st.button("✨ Apply Highlights", type="primary",
                            use_container_width=True, disabled=not uploaded_files)

with col_right:
    if not uploaded_files:
        st.markdown("""
        <div style="background:#f8fafc;border:2px dashed #cbd5e1;border-radius:12px;
                    padding:48px;text-align:center;margin-top:20px">
            <div style="font-size:48px;margin-bottom:12px">📄</div>
            <div style="font-size:18px;font-weight:600;color:#374151;margin-bottom:8px">
                No reports uploaded yet</div>
            <div style="font-size:14px;color:#6b7280;line-height:1.6">
                Upload your daily arrival PDFs on the left<br>
                then click <strong>Apply Highlights</strong><br><br>
                🟡 Yellow = normal highlight<br>
                🟠 Orange = flight with missing ETA<br>
                💬 Hover on room highlight = travelling together rooms<br>
                ✖️ Sharer bookings = no highlights at all
            </div>
        </div>""", unsafe_allow_html=True)

    elif process_btn or st.session_state.get('processed'):
        if process_btn:
            st.session_state['results'] = []
            today = date.today().strftime("%d%b").upper()
            progress_bar = st.progress(0)
            status = st.empty()

            for i, uploaded_file in enumerate(uploaded_files):
                fname = uploaded_file.name
                status.markdown(f"⏳ Processing **{fname}**...")
                progress_bar.progress(i / len(uploaded_files))
                try:
                    pdf_bytes = uploaded_file.read()
                    result_bytes, counts, total, summary_data = highlight_pdf(pdf_bytes, enabled)
                    base      = fname.replace('.pdf','').replace('.PDF','')
                    out_name  = f"{base}_{today}_highlighted.pdf"
                    html_name = f"{base}_{today}_summary.html"
                    summary_html  = build_summary_html(summary_data, pdf_filename=out_name)
                    excel_name    = f"{base}_{today}_payment_missing.xlsx"
                    excel_bytes   = build_payment_excel(summary_data)
                    st.session_state['results'].append({
                        'name': out_name, 'html_name': html_name,
                        'excel_name': excel_name,
                        'original': fname,
                        'bytes': result_bytes,
                        'html': summary_html.encode('utf-8'),
                        'excel': excel_bytes,
                        'counts': counts,
                        'total': total,
                        'summary_data': summary_data,
                        'success': True
                    })
                except Exception as e:
                    st.session_state['results'].append({
                        'name': fname, 'original': fname,
                        'error': str(e), 'success': False
                    })

            progress_bar.progress(1.0)
            status.markdown(f"✅ Done! {len(uploaded_files)} file(s) processed.")
            st.session_state['processed'] = True

        if st.session_state.get('results'):
            st.markdown("### 📥 Reports Ready")
            for r in st.session_state['results']:
                if r['success']:
                    cat_summary = " · ".join(
                        f"{CATEGORIES[k]['icon']} {v}"
                        for k, v in r['counts'].items() if v > 0)
                    st.markdown(f"""
                    <div style="background:#f0fdf4;border:1px solid #bbf7d0;
                                border-radius:8px;padding:12px 16px;margin:6px 0">
                        <strong>✅ {r['original']}</strong><br>
                        <span style="font-size:12px;color:#166534">
                            {r['total']} highlights — {cat_summary}
                        </span>
                    </div>""", unsafe_allow_html=True)
                    col_a, col_b, col_c = st.columns(3)
                    with col_a:
                        st.download_button(
                            label="⬇️ Download Highlighted PDF",
                            data=r['bytes'], file_name=r['name'],
                            mime='application/pdf', use_container_width=True,
                            key=f"dl_{r['name']}")
                    with col_b:
                        st.download_button(
                            label="🌐 Download Interactive Summary",
                            data=r['html'], file_name=r['html_name'],
                            mime='text/html', use_container_width=True,
                            key=f"html_{r['html_name']}")
                    with col_c:
                        pay_count = len([g for g in r['summary_data']['guests']
                                         if any('Payment Not Received' in f for f in g['flags'])])
                        st.download_button(
                            label=f"📊 Payment Missing ({pay_count})",
                            data=r.get('excel', b''), file_name=r.get('excel_name','payment.xlsx'),
                            mime='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                            use_container_width=True,
                            key=f"xl_{r['name']}",
                            disabled=pay_count == 0)
                    st.markdown("**Interactive Summary Preview:**")
                    st.components.v1.html(r['html'].decode('utf-8'), height=600, scrolling=True)
                else:
                    st.error(f"❌ {r['original']}: {r.get('error','Unknown error')}")
